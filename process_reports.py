"""
SAP Report Processor — Production Script
=============================================
Scans a folder for new SAP Excel P&L and Balance Sheet files,
generates clean PDFs, and updates the Aggregate P&L Output workbook.

Supports two file types:
  - PL files (Profit & Loss)  → output to "P&L Mon YY" folders
  - BS files (Balance Sheet)  → output to "Bal-Sht Mon YY" folders

Usage:
    python process_reports.py [--folder PATH] [--force] [--dry-run]

Author: Built with Claude / Anthropic
"""

import os, re, sys, json, logging, argparse
from datetime import datetime
from pathlib import Path

# ── Third-party (pip install openpyxl reportlab) ─────────────────────────────
import openpyxl
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

DEFAULT_FOLDER = os.path.join(str(Path.home()), "Documents", "AntiGravity-SAP report")

COMPANY_COLUMNS = {
    "CHALLMC":   "C",
    "CHALLPG":   "D",
    "BARWOOD":   "E",
    "TRANSITGRP":"F",
    "REG-CAB":   "G",
    "REG-FLEET": "H",
    "REG-LEASE": "I",
    "ITCURVES":  "J",
    "DISCOVERY": "K",
}

MONTH_ROWS = {
    1:  (2,  "January"),
    2:  (7,  "February"),
    3:  (12, "March"),
    4:  (22, "April"),
    5:  (27, "May"),
    6:  (32, "June"),
    7:  (42, "July"),
    8:  (47, "August"),
    9:  (52, "September"),
    10: (62, "October"),
    11: (67, "November"),
    12: (72, "December"),
}

METRIC_OFFSETS = {"revenue": 0, "payroll": 1, "other_exp": 2, "op_profit": 3, "profit": 4}

SKIP_NAMES = {
    'income - income', 'total income - income',
    'cost of goods - cost of goods sold',
    'total cost of goods - cost of goods sold',
    'total cogs - cost of goods sold',
    'general expense - general expenses',
    'total general expense - general expenses',
    'dep,disc & int - depreciation, interest & discount exp',
    'total dep,disc & int - depreciation, interest & discount exp',
    'other revenues and expenses', 'total other revenues and expenses',
    'income tax exp - income tax exp',
    'income tax - income tax expense',
    'total income tax exp - income tax exp',
    'total income tax - income tax expense',
    '#9', 'total #9', '#10', 'total #10',
}

MAJOR_TOTALS_PL = {
    'total revenues', 'total cost of sales', 'gross profit',
    'total expenses', 'operating profit', 'total financing',
    'profit after financing expenses', 'profit period',
}

# Balance Sheet major totals
MAJOR_TOTALS_BS = {
    'total assets', 'total liabilities', 'total equity',
    'total current assets - cash and equiv',
    'total fixed assets - fixed assets',
}

# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging(folder):
    log_path = os.path.join(folder, "process_log.txt")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
#  FILE DISCOVERY  (supports both PL and BS files)
# ═══════════════════════════════════════════════════════════════════════════════

MONTH_ABBREVS = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2, "FEBRURAY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}


def parse_filename(fname):
    """
    Extract (file_type, company, month_num, year) from filenames like:
        PL BARWOOD Feb 26 Annual.xlsx  →  ("PL", "BARWOOD", 2, 2026)
        BS CHALLMC JAN 26 ANNUAL.xlsx  →  ("BS", "CHALLMC", 1, 2026)
    Returns None if the file doesn't match.
    """
    stem = Path(fname).stem.upper()
    if "ANNUAL" not in stem:
        return None

    # Determine file type
    file_type = None
    if stem.startswith("PL "):
        file_type = "PL"
        core = re.sub(r'^PL\s+', '', stem)
    elif stem.startswith("BS "):
        file_type = "BS"
        core = re.sub(r'^BS\s+', '', stem)
    else:
        return None

    core = re.sub(r'\s+ANNUAL$', '', core).strip()
    tokens = core.split()

    company_tokens = []
    month_num = None
    year = None

    for i, tok in enumerate(tokens):
        if tok in MONTH_ABBREVS and month_num is None:
            month_num = MONTH_ABBREVS[tok]
            if i + 1 < len(tokens):
                yr = tokens[i + 1]
                if re.match(r'^\d{2}$', yr):
                    year = 2000 + int(yr)
            break
        else:
            company_tokens.append(tok)

    if not company_tokens or month_num is None or year is None:
        return None

    company = " ".join(company_tokens)

    matched_company = None
    for known in COMPANY_COLUMNS:
        if known.upper() in company or company in known.upper():
            matched_company = known
            break

    if matched_company is None:
        return None

    return file_type, matched_company, month_num, year


INPUT_SUBFOLDER    = "Input Files"
AGGREGATE_RELATIVE = os.path.join("Annual Financial Data", "Aggregate P&L Output.xlsx")


def find_input_files(folder, file_type_filter=None):
    """Scan the 'Input Files' subfolder (and its subfolders) for SAP Excel files.
    file_type_filter: None=all, "PL"=only P&L, "BS"=only Balance Sheet"""
    input_dir = os.path.join(folder, INPUT_SUBFOLDER)
    scan_dir  = input_dir if os.path.isdir(input_dir) else folder
    found = []

    # Walk through scan_dir and all subdirectories
    for dirpath, dirnames, filenames in os.walk(scan_dir):
        for fname in filenames:
            if not fname.lower().endswith(".xlsx"):
                continue
            result = parse_filename(fname)
            if result:
                ftype, company, month_num, year = result
                if file_type_filter and ftype != file_type_filter:
                    continue
                found.append((os.path.join(dirpath, fname), ftype, company, month_num, year))

    return sorted(found, key=lambda x: (x[4], x[3], x[1], x[2]))

# ═══════════════════════════════════════════════════════════════════════════════
#  NUMBER PARSING
# ═══════════════════════════════════════════════════════════════════════════════

def to_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("$", "").replace("\xa0", "").replace(" ", "").replace(",", "")
    # Handle percentage values
    if s.startswith("%"):
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    if not s or re.match(r'^[-_=]+$', s):
        return None
    if s == '****':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fmt_display(val):
    n = to_num(val)
    if n is None:
        return ""
    return f"({abs(n):,.2f})" if n < 0 else f"{n:,.2f}"

# ═══════════════════════════════════════════════════════════════════════════════
#  ACCOUNT NUMBER EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

ACCT_PATTERN = re.compile(r'^(\d{6}-\d{2}-\d{3}-\d{2})\s*-\s*(.+)$')

def extract_account_number(name):
    """Extract account number from name like '400000-11-000-00 - PASSENGER INCOME'.
    Returns (account_number, display_name) or (None, name)."""
    if not name:
        return None, name
    m = ACCT_PATTERN.match(name.strip())
    if m:
        return m.group(1), m.group(2).strip()
    return None, name.strip()

# ═══════════════════════════════════════════════════════════════════════════════
#  SAP EXCEL PARSING — P&L files
# ═══════════════════════════════════════════════════════════════════════════════

def is_separator_row(row):
    for v in row:
        if v is None:
            continue
        s = str(v).strip()
        if s and not re.match(r'^[\s_\-=]+$', s):
            return False
    return True


def classify_row_pl(name):
    if not name:
        return "blank"
    n = name.strip()
    nl = n.lower()
    if nl in SKIP_NAMES:
        return "skip"
    if re.match(r'^total income tax', nl):
        return "skip"
    if nl in MAJOR_TOTALS_PL:
        return "major"
    if n.startswith("Total "):
        return "subtotal"
    if ACCT_PATTERN.match(n):
        return "account"
    if n in ("Revenues", "Cost of Sales", "Expenses", "Financing", "FINANCE"):
        return "section"
    return "subheader"


def clean_display_name_pl(name, level, include_acct_num=True):
    """Clean display name. For account rows, show 'Account Number - Name'."""
    n = name.strip()
    if level == "account":
        acct_num, acct_name = extract_account_number(n)
        display = acct_name.title()
        if include_acct_num and acct_num:
            return f"{acct_num}  {display}"
        return display
    prefix = ""
    core = n
    if level == "subtotal" and n.startswith("Total "):
        prefix = "Total "
        core = n[6:]
    if " - " in core:
        left  = core.split(" - ", 1)[0].strip()
        right = core.split(" - ", 1)[1].strip().title()
        clean = right if len(right) > len(left) else left.title()
    else:
        clean = core
    return prefix + clean


def parse_sap_excel(filepath):
    """Parse a SAP Excel P&L file into cleaned display rows."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    raw, last_blank = [], True

    for row in ws.iter_rows(values_only=True):
        r = list(row) + [None, None, None]
        name, bal, ytd = r[0], r[1], r[2]

        if is_separator_row(row) or (
            name is None and to_num(bal) is None and to_num(ytd) is None
            and not str(name or "").strip()
        ):
            if not last_blank:
                raw.append(("", "", "", "blank"))
                last_blank = True
            continue

        if name is None:
            continue

        level = classify_row_pl(name)
        if level == "skip":
            continue

        if level in ("section", "subheader"):
            bal, ytd = None, None

        raw.append((clean_display_name_pl(name, level), fmt_display(bal), fmt_display(ytd), level))
        last_blank = False

    while raw and raw[-1][3] == "blank":
        raw.pop()

    pass1 = []
    for r in raw:
        if r[3] == "blank" and pass1 and pass1[-1][3] == "blank":
            continue
        pass1.append(r)

    clean = []
    for i, r in enumerate(pass1):
        if r[3] == "blank":
            prev = pass1[i - 1][3] if i > 0 else None
            nxt  = next((pass1[j][3] for j in range(i + 1, len(pass1)) if pass1[j][3] != "blank"), None)
            if prev in ("section", "subheader"):
                continue
            if nxt in ("subtotal", "major"):
                continue
        clean.append(r)

    return clean

# ═══════════════════════════════════════════════════════════════════════════════
#  SAP EXCEL PARSING — Balance Sheet files
# ═══════════════════════════════════════════════════════════════════════════════

BS_SKIP_NAMES = set()

BS_SECTION_NAMES = {
    'assets', 'liabilities', 'equity',
}

def classify_row_bs(name):
    if not name:
        return "blank"
    n = name.strip()
    nl = n.lower()
    if nl in BS_SKIP_NAMES:
        return "skip"
    if nl in MAJOR_TOTALS_BS or nl in ('total assets', 'total liabilities', 'total equity'):
        return "major"
    if n.startswith("Total "):
        return "subtotal"
    if ACCT_PATTERN.match(n):
        return "account"
    if nl in BS_SECTION_NAMES:
        return "section"
    # Sub-sections like "Current Assets - Cash and Equiv", "Fixed Assets - FIXED ASSETS"
    return "subheader"


def clean_display_name_bs(name, level, include_acct_num=True):
    """Clean display name for BS. For account rows, show 'Account Number - Name'."""
    n = name.strip()
    if level == "account":
        acct_num, acct_name = extract_account_number(n)
        # Clean non-breaking spaces
        acct_name = acct_name.replace('\xa0', ' ')
        display = acct_name.title()
        if include_acct_num and acct_num:
            return f"{acct_num}  {display}"
        return display
    prefix = ""
    core = n
    if level == "subtotal" and n.startswith("Total "):
        prefix = "Total "
        core = n[6:]
    if " - " in core:
        left  = core.split(" - ", 1)[0].strip()
        right = core.split(" - ", 1)[1].strip().title()
        clean = right if len(right) > len(left) else left.title()
    else:
        clean = core.replace('\xa0', ' ')
    return prefix + clean


def parse_bs_excel(filepath):
    """Parse a SAP Excel Balance Sheet file into cleaned display rows."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    raw, last_blank = [], True

    for row in ws.iter_rows(values_only=True):
        r = list(row) + [None, None, None, None, None]
        name = r[0]
        current_period = r[1]
        comparison_period = r[2]

        if is_separator_row(row) or (
            name is None and to_num(current_period) is None and to_num(comparison_period) is None
            and not str(name or "").strip()
        ):
            if not last_blank:
                raw.append(("", "", "", "blank"))
                last_blank = True
            continue

        if name is None:
            continue

        level = classify_row_bs(name)
        if level == "skip":
            continue

        if level in ("section", "subheader"):
            current_period, comparison_period = None, None

        raw.append((clean_display_name_bs(name, level),
                     fmt_display(current_period), fmt_display(comparison_period), level))
        last_blank = False

    while raw and raw[-1][3] == "blank":
        raw.pop()

    pass1 = []
    for r in raw:
        if r[3] == "blank" and pass1 and pass1[-1][3] == "blank":
            continue
        pass1.append(r)

    clean = []
    for i, r in enumerate(pass1):
        if r[3] == "blank":
            prev = pass1[i - 1][3] if i > 0 else None
            nxt  = next((pass1[j][3] for j in range(i + 1, len(pass1)) if pass1[j][3] != "blank"), None)
            if prev in ("section", "subheader"):
                continue
            if nxt in ("subtotal", "major"):
                continue
        clean.append(r)

    return clean

# ═══════════════════════════════════════════════════════════════════════════════
#  METRIC EXTRACTION  (for Aggregate P&L — only applies to PL files)
# ═══════════════════════════════════════════════════════════════════════════════

def extract_metrics(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    m = {"revenue": None, "payroll": None,
         "total_expenses": None, "op_profit": None, "profit": None}

    for row in ws.iter_rows(values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        bal  = to_num(row[1])
        nl   = name.lower()

        if re.search(r'^total revenues?$', nl):
            m["revenue"] = bal
        elif re.search(r'^total payroll', nl) and m["payroll"] is None:
            m["payroll"] = bal
        elif re.search(r'^total expenses?$', nl):
            m["total_expenses"] = bal
        elif re.search(r'^operating profit', nl) and m["op_profit"] is None:
            m["op_profit"] = bal
        elif re.search(r'^profit period', nl) and m["profit"] is None:
            m["profit"] = bal

    te = m["total_expenses"] or 0.0
    py = m["payroll"] or 0.0
    m["other_exp"] = round(te - py, 2) if m["total_expenses"] is not None else (0.0 if py else None)

    return m

# ═══════════════════════════════════════════════════════════════════════════════
#  PDF GENERATION — supports both PL and BS
# ═══════════════════════════════════════════════════════════════════════════════

def build_pdf(rows, company, period, update_date, output_path, report_type="P&L"):
    doc = SimpleDocTemplate(
        output_path, pagesize=letter,
        leftMargin=.75*inch, rightMargin=.75*inch,
        topMargin=.75*inch, bottomMargin=.75*inch,
    )
    S = getSampleStyleSheet()
    hB = ParagraphStyle("hB", parent=S["Normal"], alignment=TA_CENTER,
                         fontName="Helvetica-Bold", fontSize=13, spaceAfter=3)
    hN = ParagraphStyle("hN", parent=S["Normal"], alignment=TA_CENTER,
                         fontSize=10, spaceAfter=2)

    type_label = "Balance Sheet" if report_type == "BS" else "Profit & Loss"

    story = [
        Paragraph(company, hB),
        Paragraph(type_label, hN),
        Paragraph(f"PERIOD: {period}", hN),
        Paragraph(f"Update Date: {update_date}", hN),
        Spacer(1, .15 * inch),
    ]

    if report_type == "BS":
        CW = [4.0*inch, 1.5*inch, 1.5*inch]
        tdata = [["Account", "Current Period", "Comparison Period"]]
    else:
        CW = [4.0*inch, 1.5*inch, 1.5*inch]
        tdata = [["Account", "Month", "Year to Date"]]

    tstyle = [
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0),  (-1, -1), 7.5),
        ("ALIGN",         (1, 0),  (-1, -1), "RIGHT"),
        ("ALIGN",         (0, 0),  (0,  -1), "LEFT"),
        ("LINEBELOW",     (0, 0),  (-1, 0),  .8, colors.black),
        ("TOPPADDING",    (0, 0),  (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 1),
    ]

    for i, (name, col1, col2, level) in enumerate(rows):
        ri = i + 1
        if level == "blank":
            tdata.append(["", "", ""])
            continue
        indent = "        " if level == "account" else ""
        tdata.append([indent + name, col1, col2])
        if level == "major":
            tstyle += [
                ("FONTNAME",  (0, ri), (-1, ri), "Helvetica-Bold"),
                ("LINEABOVE", (1, ri), (-1, ri),  .5, colors.black),
                ("LINEBELOW", (1, ri), (-1, ri), 1.5, colors.black),
            ]
        elif level == "subtotal":
            tstyle += [
                ("FONTNAME",  (0, ri), (-1, ri), "Helvetica-Bold"),
                ("LINEABOVE", (1, ri), (-1, ri),  .5, colors.black),
            ]
        elif level == "section":
            tstyle += [
                ("FONTNAME",   (0, ri), (0,  ri), "Helvetica-Bold"),
                ("TOPPADDING", (0, ri), (-1, ri), 6),
            ]

    t = Table(tdata, colWidths=CW)
    t.setStyle(TableStyle(tstyle))
    story.append(t)
    doc.build(story)

# ═══════════════════════════════════════════════════════════════════════════════
#  AGGREGATE P&L UPDATE
# ═══════════════════════════════════════════════════════════════════════════════

NUM_FORMAT = '#,##0.00_);(#,##0.00);"-"'
MONTH_END_ROWS   = [6, 11, 16, 26, 31, 36, 46, 51, 56, 66, 71, 76]
QUARTER_END_ROWS = [21, 41, 61, 81]
YEAR_END_ROW     = 86

def _apply_bottom_border(ws, row, side, col_start=1, col_end=14):
    from openpyxl.styles import Border
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        ex = cell.border
        cell.border = Border(left=ex.left, right=ex.right, top=ex.top, bottom=side)


def apply_aggregate_formatting(ws):
    from openpyxl.styles.borders import BORDER_MEDIUM, BORDER_DOUBLE
    from openpyxl.styles import Side

    medium = Side(border_style=BORDER_MEDIUM, color="000000")
    double = Side(border_style=BORDER_DOUBLE, color="000000")

    for row in range(2, 87):
        for col in range(3, 15):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = NUM_FORMAT

    for row in MONTH_END_ROWS:
        _apply_bottom_border(ws, row, medium)
    for row in QUARTER_END_ROWS:
        _apply_bottom_border(ws, row, double)
    _apply_bottom_border(ws, YEAR_END_ROW, double)


def update_aggregate(agg_path, company, month_num, metrics, log):
    if not os.path.exists(agg_path):
        log.error(f"Aggregate file not found: {agg_path}")
        return False

    if month_num not in MONTH_ROWS:
        log.error(f"Unknown month number: {month_num}")
        return False

    col = COMPANY_COLUMNS.get(company)
    if not col:
        log.error(f"Unknown company, no column mapping: {company}")
        return False

    base_row, _ = MONTH_ROWS[month_num]

    try:
        wb = openpyxl.load_workbook(agg_path)
        ws = wb["Sheet1"]
        blue = Font(color="0000FF")

        for metric, offset in METRIC_OFFSETS.items():
            val = metrics.get(metric)
            row = base_row + offset
            cell = ws[f"{col}{row}"]
            cell.value = round(val, 2) if val is not None else 0
            cell.font  = blue

        apply_aggregate_formatting(ws)
        wb.save(agg_path)
        return True
    except PermissionError:
        log.error(f"Cannot write to Aggregate file — it may be open in Excel. Close it and retry.")
        return False
    except Exception as e:
        log.error(f"Failed to update Aggregate: {e}")
        return False

# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESSED FILES TRACKER
# ═══════════════════════════════════════════════════════════════════════════════

def load_tracker(folder):
    path = os.path.join(folder, ".processed_files.json")
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_tracker(folder, tracker):
    path = os.path.join(folder, ".processed_files.json")
    with open(path, "w") as f:
        json.dump(tracker, f, indent=2)


def file_fingerprint(filepath):
    s = os.stat(filepath)
    return f"{s.st_size}:{s.st_mtime}"

# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT FOLDER NAMING
# ═══════════════════════════════════════════════════════════════════════════════

def get_output_folder_name(file_type, short_month, yr_short):
    """Return the output folder name based on file type.
    PL → 'P&L Jan 26'
    BS → 'Bal-Sht Jan 26'
    """
    if file_type == "BS":
        return f"Bal-Sht {short_month} {yr_short}"
    else:
        return f"P&L {short_month} {yr_short}"

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="SAP Report Processor (P&L + Balance Sheet)")
    parser.add_argument("--folder",  default=DEFAULT_FOLDER, help="Report folder path")
    parser.add_argument("--force",   action="store_true",    help="Re-process already-processed files")
    parser.add_argument("--dry-run", action="store_true",    help="Show what would run without doing it")
    args = parser.parse_args()

    folder = args.folder
    if not os.path.isdir(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    log = setup_logging(folder)
    log.info("=" * 60)
    log.info("SAP Report Processor — starting (P&L + Balance Sheet)")
    log.info(f"Folder: {folder}")
    if args.dry_run:
        log.info("DRY RUN — no files will be written")

    tracker   = load_tracker(folder)
    agg_path  = os.path.join(folder, AGGREGATE_RELATIVE)
    input_files = find_input_files(folder)

    if not input_files:
        log.warning("No SAP input files found in folder.")
        return

    log.info(f"Found {len(input_files)} SAP input file(s)")

    results = {"processed": [], "skipped": [], "errors": []}

    for filepath, file_type, company, month_num, year in input_files:
        fname = os.path.basename(filepath)
        fp    = file_fingerprint(filepath)
        _, month_name = MONTH_ROWS[month_num]
        period = f"{month_name} {year}"

        if not args.force and tracker.get(fname) == fp:
            log.info(f"SKIP (unchanged): {fname}")
            results["skipped"].append(fname)
            continue

        if args.dry_run:
            log.info(f"WOULD PROCESS: {fname}  →  {file_type} / {company} / {period}")
            continue

        log.info(f"Processing: {fname}  →  {file_type} / {company} / {period}")

        try:
            short_month  = month_name[:3].title()
            yr_short     = str(year)[2:]
            month_folder_name = get_output_folder_name(file_type, short_month, yr_short)
            month_folder = os.path.join(folder, month_folder_name)
            os.makedirs(month_folder, exist_ok=True)

            update_date = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime("%m/%d/%Y %I:%M %p")
            pdf_name    = f"{company} {short_month} {yr_short}.pdf"
            pdf_path    = os.path.join(month_folder, pdf_name)

            if file_type == "BS":
                rows = parse_bs_excel(filepath)
                build_pdf(rows, company, period, update_date, pdf_path, report_type="BS")
                log.info(f"  ✓ BS PDF → {month_folder_name}/{pdf_name}")
            else:
                rows = parse_sap_excel(filepath)
                build_pdf(rows, company, period, update_date, pdf_path, report_type="PL")
                log.info(f"  ✓ PL PDF → {month_folder_name}/{pdf_name}")

                # Aggregate update only for PL files
                metrics = extract_metrics(filepath)
                ok = update_aggregate(agg_path, company, month_num, metrics, log)
                if ok:
                    log.info(f"  ✓ Aggregate updated — {company} / {period}")
                else:
                    raise RuntimeError("Aggregate update failed — see errors above")

            tracker[fname] = fp
            results["processed"].append(fname)

        except Exception as e:
            log.error(f"  ✗ FAILED: {fname} — {e}")
            results["errors"].append((fname, str(e)))

    if not args.dry_run:
        save_tracker(folder, tracker)

    log.info("-" * 60)
    log.info(f"Done.  Processed: {len(results['processed'])}  |  "
             f"Skipped (unchanged): {len(results['skipped'])}  |  "
             f"Errors: {len(results['errors'])}")
    if results["errors"]:
        log.warning("Files with errors:")
        for fname, err in results["errors"]:
            log.warning(f"  {fname}: {err}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
