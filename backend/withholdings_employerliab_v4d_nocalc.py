#!/usr/bin/env python3
"""
Extract (Batch-capable):
  1) EMPLOYEE WITHHOLDINGS
  2) EMPLOYER LIABILITIES

Adds requested changes (same as your other scripts):
✅ MULTI-PDF batch: give many PDFs -> generates one Excel per PDF automatically
✅ Option to INCLUDE/EXCLUDE MTD/QTD/YTD rows
✅ Sort rows by "Check Date" label (dates first, then MTD/QTD/YTD)
✅ Rename TOTAL columns by prefixing TABLE name:
      TOTAL      -> "<TABLE> Total"
      TOTAL (2)  -> "<TABLE> Total (2)"
✅ Add "Actual Total <TABLE>" column = sum of ALL numeric columns EXCEPT Total column(s)
✅ Add one final row at bottom that sums each numeric column (including Actual Total)

Keeps your fixes:
✅ Splits glued tokens (6,312.42ALL / 1,222.0612/05/25 / 9,689.68MTD(DEC))
✅ Prevents header band from eating first data row
✅ Keeps schema columns even if values are 0
✅ No calculations other than placing numbers into correct columns (adds only when PDF prints two numbers in same column position)

Usage:
  pip install pdfplumber pandas openpyxl numpy
  python withholdings_employerliabilities_batch_v2.py
  OR single:
  python withholdings_employerliabilities_batch_v2.py "input.pdf" "output.xlsx"
"""

import os
import re
import sys
from datetime import datetime

import pdfplumber
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# =========================
# CONFIG (hardcode here)
# =========================

# PDF_PATHS = [
#     r"MC Department summary 1.pdf",
#     r"MC Department summary 2.pdf",
#     r"MC Department summary 3.pdf",
#     r"PG Department summary 1.pdf",
#     r"PG Department summary 2.pdf",
#     r"PG Department summary 3.pdf",

# ]


# # If None -> output next to each PDF
# OUTPUT_DIR = None  # e.g. r"C:\Users\vkumawat\Downloads\Outputs"

# INCLUDE/EXCLUDE summary rows
INCLUDE_MTD = False
INCLUDE_QTD = False
INCLUDE_YTD = False

# =========================
# Regex
# =========================
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2}$")
ROWLABEL_RE = re.compile(r"^(MTD|QTD|YTD)\s*\(.*\)$|^(MTD|QTD|YTD)\(.*\)$", re.I)
NUM_RE = re.compile(r"^-?\d[\d,]*\.\d{2}$")
PCT_RE = re.compile(r"^\d+(?:\.\d+)?%$")

NUM_STATE_RE = re.compile(r"^(-?\d[\d,]*\.\d{2})([A-Za-z]{2,3})$")
NUM_DATE_RE = re.compile(r"^(-?\d[\d,]*\.\d{2})(\d{2}/\d{2}/\d{2})$")
NUM_LABEL_RE = re.compile(r"^(-?\d[\d,]*\.\d{2})((?:MTD|QTD|YTD)\(.*\))$", re.I)

X_CLUSTER_THRESHOLD = 26


def get_check_date(page):
    t = page.extract_text(layout=True) or ""
    m = re.search(r"Check\s*Date\s+(\d{2}/\d{2}/\d{2})", t, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"CheckDate\s+(\d{2}/\d{2}/\d{2})", t, flags=re.IGNORECASE)
    return m.group(1) if m else None


US_STATE_CODES = set("""
AL AK AZ AR CA CO CT DE FL GA HI ID IL IN IA KS KY LA ME MD MA MI MN MS MO MT
NE NV NH NJ NM NY NC ND OH OK OR PA RI SC SD TN TX UT VT VA WA WV WI WY DC ALL
""".split())


def to_float(s: str) -> float:
    try:
        return float(str(s).replace(",", ""))
    except Exception:
        return 0.0


def norm(s: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (s or "").upper())


def group_lines(words, y_tol=2.0):
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines, cur, cur_y = [], [], None
    for w in words:
        y = w["top"]
        if cur_y is None or abs(y - cur_y) <= y_tol:
            cur.append(w)
            cur_y = y if cur_y is None else (cur_y + y) / 2
        else:
            lines.append(sorted(cur, key=lambda x: x["x0"]))
            cur, cur_y = [w], y
    if cur:
        lines.append(sorted(cur, key=lambda x: x["x0"]))
    return lines


def normalize_header_token(t: str) -> str:
    m = {
        "1099MISC": "1099 MISC",
        "1099NEC": "1099-NEC",
        "CELLPHONE": "CELL PHONE",
        "ALLOTHER": "ALL OTHER",
        "ALLOTHERPAYMENTS": "ALL OTHER",
    }
    return m.get(t.strip().upper(), t.strip())


def is_valid_header_token(t: str) -> bool:
    if not t:
        return False
    tu = t.strip().upper()
    if tu in ("CHECK", "DATE", "MTD", "QTD", "YTD"):
        return False
    if DATE_RE.match(t) or NUM_RE.match(t) or PCT_RE.match(t):
        return False
    if len(tu) == 1 or "CONT" in tu:
        return False

    bad_exact = {
        "RUN", "RUNTIME", "PERIOD", "START", "END",
        "DEPARTMENT", "SUMMARY", "PAGE", "REPORT",
        "ORGANIZATION", "UNIT", "COMPANY", "INC",
        "CHALLENGER", "TRANSPORT", "TRANSPORTINC",
    }
    if tu in bad_exact or tu.replace(" ", "") in bad_exact:
        return False

    bad_sub = ("RUNDATE", "PERIODSTART", "PERIODEND", "CHECKDATE", "DEPARTMENTSUMMARY", "PAGE")
    n = norm(tu)
    if any(bs in n for bs in bad_sub):
        return False

    return re.search(r"[A-Z]", tu) is not None


def make_unique_in_order(names):
    counts = {}
    out = []
    for n in names:
        if n not in counts:
            counts[n] = 1
            out.append(n)
        else:
            counts[n] += 1
            out.append(f"{n} ({counts[n]})")
    return out


def detect_columns_by_x_center(header_words, x_thresh=X_CLUSTER_THRESHOLD):
    cleaned = []
    for w in header_words:
        t = (w.get("text") or "").strip()
        if not is_valid_header_token(t):
            continue
        x_center = (w["x0"] + w["x1"]) / 2
        cleaned.append({**w, "x_center": x_center, "text": normalize_header_token(t)})

    if not cleaned:
        return []

    cleaned = sorted(cleaned, key=lambda x: x["x_center"])
    clusters, cur = [], [cleaned[0]]
    for w in cleaned[1:]:
        if abs(w["x_center"] - cur[-1]["x_center"]) <= x_thresh:
            cur.append(w)
        else:
            clusters.append(cur)
            cur = [w]
    clusters.append(cur)

    cols = []
    for c in clusters:
        c_sorted = sorted(c, key=lambda x: (x["top"], x["x0"]))
        title = " ".join([x["text"] for x in c_sorted])
        title = " ".join(title.split()).strip()
        mid = sum(x["x_center"] for x in c) / len(c)
        cols.append((title, mid))
    return cols


def find_headers_on_page(page):
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    lines = group_lines(words)
    headers = []

    for ln in lines:
        line = " ".join([w["text"] for w in ln]).strip()
        n = norm(line)
        if "EMPLOYEEWITHHOLDINGS" in n:
            headers.append(("EMPLOYEE WITHHOLDINGS", ln[0]["top"]))
        if "EMPLOYERLIABILITIES" in n:
            headers.append(("EMPLOYER LIABILITIES", ln[0]["top"]))
        if "EMPLOYEEDEDUCTIONS" in n:
            headers.append(("EMPLOYEE DEDUCTIONS", ln[0]["top"]))
        if "EMPLOYEELIABILITIES" in n:
            headers.append(("EMPLOYEE LIABILITIES", ln[0]["top"]))

    headers = sorted(headers, key=lambda x: x[1])
    cleaned = []
    for name, y in headers:
        if not cleaned or abs(y - cleaned[-1][1]) > 5 or name != cleaned[-1][0]:
            cleaned.append((name, y))
    return cleaned


def compute_section_regions(pdf, target_name):
    regions = []
    started = False
    for p_idx, page in enumerate(pdf.pages):
        hs = find_headers_on_page(page)
        y_target = None
        y_stop = None

        for name, y in hs:
            if name == target_name and y_target is None:
                y_target = y

        if not started:
            if y_target is None:
                continue
            started = True
            y0 = y_target - 2
            for name, y in hs:
                if y > y_target and name != target_name:
                    y_stop = y
                    break
        else:
            y0 = 0
            for name, y in hs:
                if name != target_name:
                    y_stop = y
                    break

        y1 = page.height - 5 if y_stop is None else y_stop - 2
        regions.append((p_idx, y0, y1))
        if y_stop is not None:
            break
    return regions


def split_special_words(words):
    """Split glued tokens like:
      - 928.72DC  -> 928.72 + DC
      - 6,312.42ALL -> 6,312.42 + ALL
      - 1,222.0612/05/25 -> 1,222.06 + 12/05/25
      - 9,689.68MTD(DEC) -> 9,689.68 + MTD(DEC)
    """
    out = []
    for w in words:
        t = (w.get("text") or "").strip()

        m = NUM_DATE_RE.match(t)
        if m:
            num, d = m.group(1), m.group(2)
            x0, x1 = w["x0"], w["x1"]
            mid = x0 + (x1 - x0) * 0.62
            out.append({**w, "text": num, "x1": mid})
            out.append({**w, "text": d, "x0": mid})
            continue

        m = NUM_LABEL_RE.match(t)
        if m:
            num, lab = m.group(1), m.group(2).upper()
            x0, x1 = w["x0"], w["x1"]
            mid = x0 + (x1 - x0) * 0.68
            out.append({**w, "text": num, "x1": mid})
            out.append({**w, "text": lab, "x0": mid})
            continue

        m = NUM_STATE_RE.match(t)
        if m and m.group(2).upper() in US_STATE_CODES:
            num, st = m.group(1), m.group(2).upper()
            x0, x1 = w["x0"], w["x1"]
            mid = x0 + (x1 - x0) * 0.78
            out.append({**w, "text": num, "x1": mid})
            out.append({**w, "text": st, "x0": mid})
            continue

        out.append(w)
    return out


def compute_header_band(lines, seg, y0, ci, di):
    check_ln = lines[ci]
    date_ln = lines[di]
    header_top = max(y0, min(w["top"] for w in check_ln) - 20)
    base_bottom = max(w["bottom"] for w in date_ln) + 3

    # find first data row so we NEVER include it inside the header band
    first_data_top = None
    for j in range(di + 1, min(di + 40, len(lines))):
        if not lines[j]:
            continue
        first = lines[j][0]["text"].strip()
        fu = first.upper()
        if DATE_RE.match(first) or ROWLABEL_RE.match(fu) or ROWLABEL_RE.match(fu.replace(" ", "")):
            first_data_top = lines[j][0]["top"]
            break

    header_bottom = base_bottom
    cap = (first_data_top - 0.5) if first_data_top is not None else None

    # extend for wrapped header fragments but do NOT cross into the first data row
    for j in range(di + 1, min(di + 10, len(lines))):
        if not lines[j]:
            continue
        first = lines[j][0]["text"].strip()
        fu = first.upper()
        if DATE_RE.match(first) or ROWLABEL_RE.match(fu) or ROWLABEL_RE.match(fu.replace(" ", "")):
            break
        header_bottom = max(header_bottom, max(w["bottom"] for w in lines[j]) + 1)
        if cap is not None:
            header_bottom = min(header_bottom, cap)

    header_band = [w for w in seg if header_top <= w["top"] <= header_bottom]
    return header_top, header_bottom, header_band


def extract_subtables(page, page_idx, y0, y1):
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    words = split_special_words(words)

    seg = [w for w in words if y0 <= w["top"] <= y1]
    if not seg:
        return []

    lines = group_lines(seg)

    # find CHECK / DATE header pairs
    check_pairs = []
    for i, ln in enumerate(lines):
        if ln and ln[0]["text"].strip().upper() == "CHECK":
            for j in range(i + 1, min(i + 12, len(lines))):
                if lines[j] and lines[j][0]["text"].strip().upper() == "DATE":
                    check_pairs.append((i, j))
                    break

    subtables = []
    for (ci, di) in check_pairs:
        _, header_bottom, header_band = compute_header_band(lines, seg, y0, ci, di)

        cols = detect_columns_by_x_center(header_band)
        if not cols:
            continue

        col_names = make_unique_in_order([c[0] for c in cols])
        col_mids = {col_names[i]: cols[i][1] for i in range(len(cols))}

        row_y0 = header_bottom + 0.5

        # row_y1 until next CHECK (if any) otherwise section end
        next_check_top = None
        for (ci2, _) in check_pairs:
            if lines[ci2][0]["top"] > lines[ci][0]["top"]:
                next_check_top = lines[ci2][0]["top"]
                break
        row_y1 = (next_check_top - 1) if next_check_top is not None else y1

        rows = []
        for ln in lines:
            if not ln:
                continue
            if ln[0]["top"] < row_y0 or ln[0]["top"] > row_y1:
                continue

            first = ln[0]["text"].strip()
            fu = first.upper()

            if DATE_RE.match(first) or ROWLABEL_RE.match(fu) or ROWLABEL_RE.match(fu.replace(" ", "")):
                label = first
                # If the first token isn't a valid date (e.g. MTD), try page-level date as fallback
                if not DATE_RE.match(label):
                    cd_fallback = get_check_date(page)
                    if cd_fallback:
                        label = cd_fallback

                row = {"Check Date": label, "_page": page_idx, "_table_top": float(lines[ci][0]["top"]), "States": ""}
                for c in col_mids:
                    row[c] = 0.0

                for w in ln[1:]:
                    txt = (w.get("text") or "").strip()
                    if not txt:
                        continue

                    if txt.upper() in US_STATE_CODES:
                        st = txt.upper()
                        row["States"] = (row["States"] + " " + st).strip() if row["States"] else st
                        continue

                    if PCT_RE.match(txt):
                        x_mid = (w["x0"] + w["x1"]) / 2
                        closest = min(col_mids.items(), key=lambda kv: abs(x_mid - kv[1]))[0]
                        row[closest] = txt
                        continue

                    if NUM_RE.match(txt):
                        val = to_float(txt)
                        x_mid = (w["x0"] + w["x1"]) / 2
                        closest = min(col_mids.items(), key=lambda kv: abs(x_mid - kv[1]))[0]
                        row[closest] = float(row.get(closest, 0.0)) + val

                rows.append(row)

        if rows:
            df = pd.DataFrame(rows).fillna(0.0)
            subtables.append(df)

    subtables.sort(key=lambda d: (int(d["_page"].iloc[0]), float(d["_table_top"].iloc[0])))
    return subtables


def add_occur(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Occur"] = df.groupby("Check Date").cumcount() + 1
    return df


def merge_on_date_occur(master: pd.DataFrame, part: pd.DataFrame) -> pd.DataFrame:
    key_cols = ["Check Date", "Occur"]
    m = master.copy()
    p = part.copy()
    if "Occur" not in m.columns:
        m = add_occur(m)
    if "Occur" not in p.columns:
        p = add_occur(p)

    bring = [c for c in p.columns if c not in key_cols and not c.startswith("_")]
    merged = pd.merge(m, p[key_cols + bring], on=key_cols, how="left", suffixes=("", "__dup")).fillna(0.0)

    for c in list(merged.columns):
        if c.endswith("__dup"):
            base = c[:-5].strip()
            if base in merged.columns:
                a = pd.to_numeric(merged[base], errors="coerce")
                b = pd.to_numeric(merged[c], errors="coerce")
                if a.notna().any() and b.notna().any():
                    merged[base] = a.fillna(0.0) + b.fillna(0.0)
                else:
                    merged[base] = (
                        merged[base].astype(str).replace("0.0", "").str.strip()
                        + " "
                        + merged[c].astype(str).replace("0.0", "").str.strip()
                    ).str.strip()
                merged.drop(columns=[c], inplace=True)
            else:
                merged.rename(columns={c: base}, inplace=True)

    return merged


# =========================
# NEW: filtering/sorting/totals (same as your other scripts)
# =========================
def filter_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if not INCLUDE_MTD:
        out = out[~out["Check Date"].astype(str).str.upper().str.startswith("MTD", na=False)]
    if not INCLUDE_QTD:
        out = out[~out["Check Date"].astype(str).str.upper().str.startswith("QTD", na=False)]
    if not INCLUDE_YTD:
        out = out[~out["Check Date"].astype(str).str.upper().str.startswith("YTD", na=False)]
    return out.reset_index(drop=True)


def sort_key_label(label: str):
    s = (label or "").strip().upper()
    if DATE_RE.match(s):
        try:
            return (0, datetime.strptime(s, "%m/%d/%y"))
        except Exception:
            return (0, datetime.max)
    if s.startswith("MTD"):
        return (1, s)
    if s.startswith("QTD"):
        return (2, s)
    if s.startswith("YTD"):
        return (3, s)
    return (4, s)


def sort_by_check_date_label(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["_k"] = out["Check Date"].apply(sort_key_label)
    out = out.sort_values(["_k", "Occur"], kind="mergesort").drop(columns=["_k"]).reset_index(drop=True)
    return out


def rename_total_cols_with_prefix(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    """
    TOTAL -> "<table> Total"
    TOTAL (2) -> "<table> Total (2)"
    """
    out = df.copy()
    ren = {}
    for c in out.columns:
        cu = str(c).strip().upper()
        if cu == "TOTAL" or cu.startswith("TOTAL ("):
            suffix = "" if cu == "TOTAL" else c[len("TOTAL"):]
            ren[c] = f"{table_name} Total{suffix}"
    return out.rename(columns=ren)


def add_actual_total(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    """
    Adds: "Actual Total <table_name>" = sum of ALL numeric columns EXCEPT the Total column(s).
    """
    out = df.copy()
    key_cols = {"Check Date", "Occur", "States"}
    total_prefix = f"{table_name.strip().upper()} TOTAL"

    def is_total_col(col: str) -> bool:
        up = str(col).strip().upper()
        return up == total_prefix or up.startswith(total_prefix + " (")

    numeric_cols = [c for c in out.columns if c not in key_cols and not is_total_col(c)]

    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    out[f"Actual Total {table_name.title()}"] = out[numeric_cols].sum(axis=1)
    return out


def add_bottom_sum_row(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    sum_row = {c: "" for c in out.columns}
    sum_row["Check Date"] = "TOTAL"
    sum_row["Occur"] = ""
    sum_row["States"] = ""
    for c in out.columns:
        if c in ("Check Date", "Occur", "States"):
            continue
        sum_row[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).sum()
    return pd.concat([out, pd.DataFrame([sum_row])], ignore_index=True)


def build_table(pdf_path: str, section_name: str) -> pd.DataFrame:
    with pdfplumber.open(pdf_path) as pdf:
        regions = compute_section_regions(pdf, section_name)
        subtables = []
        for (p_idx, y0, y1) in regions:
            subtables.extend(extract_subtables(pdf.pages[p_idx], p_idx, y0, y1))

    if not subtables:
        return pd.DataFrame()

    def _row_sort_key(label: str):
        lab = str(label).strip()
        if DATE_RE.match(lab):
            try:
                mm, dd, yy = lab.split("/")
                return (0, int(yy), int(mm), int(dd), 0)
            except Exception:
                return (0, 99, 99, 99, 0)
        u = lab.upper().replace(" ", "")
        if u.startswith("MTD"):
            return (1, 0, 0, 0, 1)
        if u.startswith("QTD"):
            return (1, 0, 0, 0, 2)
        if u.startswith("YTD"):
            return (1, 0, 0, 0, 3)
        return (2, 0, 0, 0, 9)

    # UNION of labels
    all_labels = []
    for t in subtables:
        for lab in t["Check Date"].tolist():
            if lab not in all_labels:
                all_labels.append(lab)

    master = pd.DataFrame({"Check Date": sorted(all_labels, key=_row_sort_key)})
    master = add_occur(master).fillna(0.0)

    col_order = list(master.columns)
    for part in subtables:
        part_clean = part[[c for c in part.columns if not c.startswith("_")]]
        part_clean = add_occur(part_clean).fillna(0.0)
        master = merge_on_date_occur(master, part_clean)
        for c in master.columns:
            if c not in col_order:
                col_order.append(c)

    rest = [c for c in col_order if c not in ("Check Date", "Occur")]
    df = master[["Check Date", "Occur"] + rest].fillna(0.0)

    # ===== apply requested transformations =====
    df = filter_summary_rows(df)
    df = sort_by_check_date_label(df)

    # numeric coercion for non-identifiers (keep States text)
    for c in df.columns:
        if c in ("Check Date", "Occur", "States"):
            continue
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    # rename Total columns with prefix
    df = rename_total_cols_with_prefix(df, section_name)

    # add actual total excluding Total cols
    df = add_actual_total(df, section_name)

    # bottom sum row
    df = add_bottom_sum_row(df)

    return df


def write_workbook(out_path: str, tables: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for sheet_name, df in tables.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df is None or df.empty:
            ws.append(["No data found"])
            continue

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # format numeric columns (skip id cols + States)
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=ws.max_column):
            header = (col[0].value or "")
            if header in ("Check Date", "Occur", "States"):
                continue
            for cell in col:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "A2"

    wb.save(out_path)


def make_output_path(pdf_path: str) -> str:
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    out_name = f"{base}_Withholdings_EmployerLiabilities.xlsx"
    if OUTPUT_DIR:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        return os.path.join(OUTPUT_DIR, out_name)
    return os.path.join(os.path.dirname(pdf_path) or ".", out_name)


def run_one(pdf_path: str, out_path: str):
    tables = {
        "Employee Withholdings": build_table(pdf_path, "EMPLOYEE WITHHOLDINGS"),
        "Employer Liabilities": build_table(pdf_path, "EMPLOYER LIABILITIES"),
    }
    write_workbook(out_path, tables)
    print(f"✅ Done: {out_path}")


def main():
    ok = 0
    bad = 0
    for pdf_path in PDF_PATHS:
        if not os.path.exists(pdf_path):
            print("❌ Missing PDF:", pdf_path)
            bad += 1
            continue
        out_path = make_output_path(pdf_path)
        try:
            run_one(pdf_path, out_path)
            ok += 1
        except Exception as e:
            print(f"❌ ERROR for {pdf_path}: {type(e).__name__}: {e}")
            bad += 1

    print(f"Batch complete. Success={ok} Failed={bad}")


if __name__ == "__main__":
    # Optional single-file CLI:
    #   python script.py "input.pdf" "output.xlsx"
    if len(sys.argv) >= 3:
        run_one(sys.argv[1], sys.argv[2])
    else:
        main()
