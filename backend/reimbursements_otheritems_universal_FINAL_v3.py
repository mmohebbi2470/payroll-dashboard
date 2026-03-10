# reimbursements_otheritems_universal_FINAL_BATCH_v2.py
# ✅ Extracts:
#   1) REIMBURSEMENTS & OTHER PAYMENTS
#   2) OTHER ITEMS (DO NOT INCREASE NET PAY)
#
# ✅ Batch mode: multiple PDFs -> one Excel per PDF (2 sheets)
#
# Adds (same changes you asked earlier):
# ✅ Option to INCLUDE/EXCLUDE MTD/QTD/YTD rows
# ✅ Sort rows by "Check Date" label (dates first, then MTD/QTD/YTD)
# ✅ Rename TOTAL columns by prefixing the TABLE name:
#       TOTAL      -> "<TABLE> Total"
#       TOTAL (2)  -> "<TABLE> Total (2)"
# ✅ Add "Actual Total <TABLE>" column = sum of ALL numeric columns EXCEPT Total column(s)
# ✅ Add one final row at bottom that sums each numeric column (including Actual Total)
#
# Usage:
#   pip install pdfplumber pandas openpyxl numpy
#   python reimbursements_otheritems_universal_FINAL_BATCH_v2.py
#   OR: python script.py "one.pdf" "one.xlsx"

import os
import sys
import re
from datetime import datetime

import pdfplumber
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill


# =========================
# CONFIG (hardcode here)
# =========================

# MULTI PDF INPUTS (hardcode all PDFs here)
# PDF_PATHS = [
#     r"MC Department summary 1.pdf",
#     r"MC Department summary 2.pdf",
#     r"MC Department summary 3.pdf",
#     r"PG Department summary 1.pdf",
#     r"PG Department summary 2.pdf",
#     r"PG Department summary 3.pdf",

# ]

# # If None -> output next to each PDF
# OUTPUT_DIR = None  # e.g. r"C:\Users\vkumawat\Downloads\Outputs"

# INCLUDE/EXCLUDE summary rows
INCLUDE_MTD = False
INCLUDE_QTD = False
INCLUDE_YTD = False


# =========================
# Regex
# =========================
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2}$")
ROWLABEL_RE = re.compile(r"^(MTD|QTD|YTD)\s*\(.*\)$|^(MTD|QTD|YTD)\(.*\)$", re.I)
NUM_RE = re.compile(r"^-?\d[\d,]*\.\d{2}$")
X_CLUSTER_THRESHOLD = 26


def to_float(s: str) -> float:
    try:
        return float(str(s).replace(",", ""))
    except Exception:
        return 0.0


def norm(s: str) -> str:
    # normalize for robust header matching (spaces/punctuation/wrapping)
    return re.sub(r"[^A-Z0-9]+", "", (s or "").upper())


def group_lines(words, y_tol=2.0):
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines, cur, cur_y = [], [], None
    for w in words:
        y = w["top"]
        if cur_y is None or abs(y - cur_y) <= y_tol:
            cur.append(w)
            cur_y = y if cur_y is None else (cur_y + y) / 2
        else:
            lines.append(sorted(cur, key=lambda x: x["x0"]))
            cur, cur_y = [w], y
    if cur:
        lines.append(sorted(cur, key=lambda x: x["x0"]))
    return lines


def normalize_header_token(t: str) -> str:
    m = {
        "1099MISC": "1099 MISC",
        "CELLPHONE": "CELL PHONE",
        "ALLOTHER": "ALL OTHER",
        "ALLOTHERPAYMENTS": "ALL OTHER",
    }
    return m.get(t.strip().upper(), t.strip())


def is_valid_header_token(t: str) -> bool:
    if not t:
        return False
    tu = t.strip().upper()

    if tu in ("CHECK", "DATE", "MTD", "QTD", "YTD"):
        return False
    if DATE_RE.match(t) or NUM_RE.match(t):
        return False
    if len(tu) == 1:
        return False
    if "CONT" in tu:
        return False

    bad_exact = {
        "RUN", "RUNTIME", "PERIOD", "START", "END",
        "DEPARTMENT", "SUMMARY", "PAGE", "REPORT",
        "ORGANIZATION", "UNIT", "COMPANY", "INC",
        "CHALLENGER", "TRANSPORT", "TRANSPORTINC",
    }
    if tu in bad_exact or tu.replace(" ", "") in bad_exact:
        return False

    bad_sub = ("RUNDATE", "PERIODSTART", "PERIODEND", "CHECKDATE", "DEPARTMENTSUMMARY", "PAGE")
    n = norm(tu)
    if any(bs in n for bs in bad_sub):
        return False

    if "REIMBURSEMENTS" in tu or "OTHERITEMS" in tu.replace(" ", ""):
        return False

    return re.search(r"[A-Z]", tu) is not None


def make_unique_in_order(names):
    counts = {}
    out = []
    for n in names:
        if n not in counts:
            counts[n] = 1
            out.append(n)
        else:
            counts[n] += 1
            out.append(f"{n} ({counts[n]})")
    return out


def detect_columns_by_x_center(header_words, x_thresh=X_CLUSTER_THRESHOLD):
    cleaned = []
    for w in header_words:
        t = (w.get("text") or "").strip()
        if not is_valid_header_token(t):
            continue
        x_center = (w["x0"] + w["x1"]) / 2
        cleaned.append({**w, "x_center": x_center, "text": normalize_header_token(t)})

    if not cleaned:
        return []

    cleaned = sorted(cleaned, key=lambda x: x["x_center"])
    clusters, cur = [], [cleaned[0]]
    for w in cleaned[1:]:
        if abs(w["x_center"] - cur[-1]["x_center"]) <= x_thresh:
            cur.append(w)
        else:
            clusters.append(cur)
            cur = [w]
    clusters.append(cur)

    cols = []
    for c in clusters:
        c_sorted = sorted(c, key=lambda x: (x["top"], x["x0"]))
        title = " ".join([x["text"] for x in c_sorted])
        title = " ".join(title.split()).strip()
        mid = sum(x["x_center"] for x in c) / len(c)
        cols.append((title, mid))
    return cols


def find_headers_on_page(page):
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    lines = group_lines(words)
    headers = []

    for ln in lines:
        line = " ".join([w["text"] for w in ln]).strip()
        n = norm(line)

        # Strict Reimbursements header
        if "REIMBURSEMENTS" in n and "OTHERPAYMENTS" in n and "ALLOTHER" not in n:
            headers.append(("REIMBURSEMENTS & OTHER PAYMENTS", ln[0]["top"]))

        # Strict Other Items header (must include the full unique phrase)
        if n.startswith("OTHERITEMS") and "DONOTINCREASENETPAY" in n:
            headers.append(("OTHER ITEMS", ln[0]["top"]))

        # Hard stops
        if "EMPLOYEEWITHHOLDINGS" in n:
            headers.append(("EMPLOYEE WITHHOLDINGS", ln[0]["top"]))
        if "EMPLOYEEDEDUCTIONS" in n:
            headers.append(("EMPLOYEE DEDUCTIONS", ln[0]["top"]))
        if "EMPLOYEELIABILITIES" in n:
            headers.append(("EMPLOYEE LIABILITIES", ln[0]["top"]))
        if "EMPLOYERLIABILITIES" in n:
            headers.append(("EMPLOYER LIABILITIES", ln[0]["top"]))

    headers = sorted(headers, key=lambda x: x[1])
    cleaned = []
    for name, y in headers:
        if not cleaned or abs(y - cleaned[-1][1]) > 5 or name != cleaned[-1][0]:
            cleaned.append((name, y))
    return cleaned


def compute_section_regions(pdf, target_name):
    regions = []
    started = False
    for p_idx, page in enumerate(pdf.pages):
        hs = find_headers_on_page(page)
        y_target = None
        y_stop = None

        for name, y in hs:
            if name == target_name and y_target is None:
                y_target = y

        if not started:
            if y_target is None:
                continue
            started = True
            y0 = y_target - 2
            for name, y in hs:
                if y > y_target and name != target_name:
                    y_stop = y
                    break
        else:
            y0 = 0
            for name, y in hs:
                if name != target_name:
                    y_stop = y
                    break

        y1 = page.height - 5 if y_stop is None else y_stop - 2
        regions.append((p_idx, y0, y1))
        if y_stop is not None:
            break
    return regions


def extract_subtables(page, page_idx, y0, y1):
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    seg = [w for w in words if y0 <= w["top"] <= y1]
    if not seg:
        return []

    lines = group_lines(seg)

    check_idxs = []
    for i, ln in enumerate(lines):
        if ln and ln[0]["text"].strip().upper() == "CHECK":
            ok = False
            for j in range(i + 1, min(i + 12, len(lines))):
                if lines[j] and lines[j][0]["text"].strip().upper() == "DATE":
                    ok = True
                    break
            if ok:
                check_idxs.append(i)

    subtables = []
    for k, ci in enumerate(check_idxs):
        check_ln = lines[ci]
        date_ln = None
        for j in range(ci + 1, min(ci + 12, len(lines))):
            if lines[j] and lines[j][0]["text"].strip().upper() == "DATE":
                date_ln = lines[j]
                break
        if date_ln is None:
            continue

        header_top = max(y0, min(w["top"] for w in check_ln) - 20)
        header_bottom = max(w["bottom"] for w in date_ln) + 3
        header_band = [w for w in seg if header_top <= w["top"] <= header_bottom]

        cols = detect_columns_by_x_center(header_band)
        if not cols:
            continue
        col_names = make_unique_in_order([c[0] for c in cols])
        col_mids = {col_names[i]: cols[i][1] for i in range(len(cols))}

        row_y0 = header_bottom + 1
        next_ci = check_idxs[k + 1] if k + 1 < len(check_idxs) else None
        row_y1 = y1
        if next_ci is not None:
            row_y1 = min(row_y1, lines[next_ci][0]["top"] - 2)

        rows = []
        for ln in lines:
            if not ln:
                continue
            if ln[0]["top"] < row_y0 or ln[0]["top"] > row_y1:
                continue
            first = ln[0]["text"].strip()
            fu = first.upper()
            is_date = DATE_RE.match(first)
            is_summary = ROWLABEL_RE.match(fu.replace(" ", "")) or ROWLABEL_RE.match(fu)

            # Skip MTD/QTD/YTD rows entirely — only keep actual
            # payroll-date rows. This prevents page-break continuations
            # from duplicating data when summary rows appear on the next page.
            if is_summary:
                continue

            if is_date:
                row = {"Check Date": first, "_page": page_idx, "_table_top": float(check_ln[0]["top"])}
                for c in col_mids:
                    row[c] = 0.0
                nums = [w for w in ln[1:] if NUM_RE.match(w["text"].strip())]
                for nword in nums:
                    val = to_float(nword["text"].strip())
                    x_mid = (nword["x0"] + nword["x1"]) / 2
                    closest = min(col_mids.items(), key=lambda kv: abs(x_mid - kv[1]))[0]
                    row[closest] += val
                rows.append(row)

        if not rows:
            continue
        df = pd.DataFrame(rows).fillna(0.0)
        subtables.append(df)

    subtables.sort(key=lambda d: (int(d["_page"].iloc[0]), float(d["_table_top"].iloc[0])))
    return subtables


def add_occur(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Occur"] = df.groupby("Check Date").cumcount() + 1
    return df


def merge_on_date_occur(master: pd.DataFrame, part: pd.DataFrame) -> pd.DataFrame:
    key_cols = ["Check Date", "Occur"]
    m = master.copy()
    p = part.copy()
    if "Occur" not in m.columns:
        m = add_occur(m)
    if "Occur" not in p.columns:
        p = add_occur(p)

    bring = [c for c in p.columns if c not in key_cols and not c.startswith("_")]
    merged = pd.merge(m, p[key_cols + bring], on=key_cols, how="left", suffixes=("", "__dup")).fillna(0.0)

    for c in list(merged.columns):
        if c.endswith("__dup"):
            base = c[:-5].strip()
            if base in merged.columns:
                merged[base] = (
                    pd.to_numeric(merged[base], errors="coerce").fillna(0.0)
                    + pd.to_numeric(merged[c], errors="coerce").fillna(0.0)
                )
                merged.drop(columns=[c], inplace=True)
            else:
                k = 2
                new = f"{base} ({k})"
                while new in merged.columns:
                    k += 1
                    new = f"{base} ({k})"
                merged.rename(columns={c: new}, inplace=True)

    return merged


# =========================
# NEW: filtering/sorting/totals (same as you asked)
# =========================
def filter_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if not INCLUDE_MTD:
        out = out[~out["Check Date"].astype(str).str.upper().str.startswith("MTD", na=False)]
    if not INCLUDE_QTD:
        out = out[~out["Check Date"].astype(str).str.upper().str.startswith("QTD", na=False)]
    if not INCLUDE_YTD:
        out = out[~out["Check Date"].astype(str).str.upper().str.startswith("YTD", na=False)]
    return out.reset_index(drop=True)


def sort_key_label(label: str):
    s = (label or "").strip().upper()
    if DATE_RE.match(s):
        try:
            return (0, datetime.strptime(s, "%m/%d/%y"))
        except Exception:
            return (0, datetime.max)
    if s.startswith("MTD"):
        return (1, s)
    if s.startswith("QTD"):
        return (2, s)
    if s.startswith("YTD"):
        return (3, s)
    return (4, s)


def sort_by_check_date_label(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["_k"] = out["Check Date"].apply(sort_key_label)
    out = out.sort_values(["_k", "Occur"], kind="mergesort").drop(columns=["_k"]).reset_index(drop=True)
    return out


def rename_total_cols_with_prefix(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    """
    TOTAL -> "<table> Total"
    TOTAL (2) -> "<table> Total (2)"
    """
    out = df.copy()
    ren = {}
    for c in out.columns:
        cu = str(c).strip().upper()
        if cu == "TOTAL" or cu.startswith("TOTAL ("):
            suffix = "" if cu == "TOTAL" else c[len("TOTAL"):]  # keeps " (2)" etc
            ren[c] = f"{table_name} Total{suffix}"
    return out.rename(columns=ren)


def add_actual_total(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    """
    Adds: "Actual Total <table_name>" = sum of ALL numeric columns EXCEPT the Total column(s).
    """
    out = df.copy()
    key_cols = {"Check Date", "Occur"}
    total_prefix = f"{table_name.strip().upper()} TOTAL"

    def is_total_col(col: str) -> bool:
        up = str(col).strip().upper()
        return up == total_prefix or up.startswith(total_prefix + " (")

    numeric_cols = [c for c in out.columns if c not in key_cols and not is_total_col(c)]

    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    out[f"Actual Total {table_name}"] = out[numeric_cols].sum(axis=1)
    return out


def add_bottom_sum_row(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    sum_row = {c: "" for c in out.columns}
    sum_row["Check Date"] = "TOTAL"
    sum_row["Occur"] = ""
    for c in out.columns:
        if c in ("Check Date", "Occur"):
            continue
        sum_row[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).sum()
    return pd.concat([out, pd.DataFrame([sum_row])], ignore_index=True)


# =========================
# Build table (one section)
# =========================
def build_table(pdf_path: str, section_name: str) -> pd.DataFrame:
    with pdfplumber.open(pdf_path) as pdf:
        regions = compute_section_regions(pdf, section_name)
        subtables = []
        for (p_idx, y0, y1) in regions:
            subtables.extend(extract_subtables(pdf.pages[p_idx], p_idx, y0, y1))

    if not subtables:
        return pd.DataFrame()

    # ✅ Build GLOBAL master row index as UNION of labels (handles split dates across pages)
    def _row_sort_key(label: str):
        lab = str(label).strip()
        if DATE_RE.match(lab):
            try:
                mm, dd, yy = lab.split("/")
                return (0, int(yy), int(mm), int(dd), 0)
            except Exception:
                return (0, 99, 99, 99, 0)
        u = lab.upper().replace(" ", "")
        if u.startswith("MTD"):
            return (1, 0, 0, 0, 1)
        if u.startswith("QTD"):
            return (1, 0, 0, 0, 2)
        if u.startswith("YTD"):
            return (1, 0, 0, 0, 3)
        return (2, 0, 0, 0, 9)

    all_labels = []
    for t in subtables:
        for lab in t["Check Date"].tolist():
            if lab not in all_labels:
                all_labels.append(lab)

    all_labels_sorted = sorted(all_labels, key=_row_sort_key)

    master = pd.DataFrame({"Check Date": all_labels_sorted})
    master = add_occur(master).fillna(0.0)

    col_order = list(master.columns)
    parts = subtables[:]
    parts.sort(key=lambda d: (int(d["_page"].iloc[0]), float(d["_table_top"].iloc[0])))

    for part in parts:
        part_clean = part[[c for c in part.columns if not c.startswith("_")]]
        part_clean = part_clean.drop_duplicates().reset_index(drop=True)
        part_clean = add_occur(part_clean).fillna(0.0)
        master = merge_on_date_occur(master, part_clean)
        for c in master.columns:
            if c not in col_order:
                col_order.append(c)

    # ✅ Safety: OTHER ITEMS should never look like Employee Withholdings.
    if section_name == "OTHER ITEMS":
        bad_cols = {"SOCSEC", "MEDICARE", "FEDINCOME", "STATEINCOME", "LOCALINCOME", "ALLOTHERWITHHOLDINGS", "EMPLOYEEWITHHOLDINGS"}
        for c in master.columns:
            cn = norm(str(c))
            if any(b in cn for b in bad_cols):
                return pd.DataFrame()

    rest = [c for c in col_order if c not in ("Check Date", "Occur")]
    df = master[["Check Date", "Occur"] + rest].fillna(0.0)

    # ===== apply requested transformations =====
    df = filter_summary_rows(df)
    df = sort_by_check_date_label(df)

    # numeric + fill
    for c in df.columns:
        if c not in ("Check Date", "Occur"):
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    # rename TOTAL columns using table prefix
    df = rename_total_cols_with_prefix(df, section_name)

    # add actual total excluding total cols
    df = add_actual_total(df, section_name)

    # bottom sum row
    df = add_bottom_sum_row(df)

    return df


# =========================
# Write workbook (two sheets)
# =========================
def write_workbook(out_path: str, tables: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for sheet_name, df in tables.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df is None or df.empty:
            ws.append(["No data found"])
            continue

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if ws.max_column >= 3:
            for col in ws.iter_cols(min_row=2, min_col=3, max_col=ws.max_column):
                for cell in col:
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "A2"

    wb.save(out_path)


def make_output_path(pdf_path: str) -> str:
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    out_name = f"{base}_Reimb_OtherItems.xlsx"
    if OUTPUT_DIR:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        return os.path.join(OUTPUT_DIR, out_name)
    return os.path.join(os.path.dirname(pdf_path) or ".", out_name)


def run_one(pdf_path: str, out_path: str):
    tables = {
        "Reimbursements & Other Payments": build_table(pdf_path, "REIMBURSEMENTS & OTHER PAYMENTS"),
        "Other Items": build_table(pdf_path, "OTHER ITEMS"),
    }
    write_workbook(out_path, tables)
    print(f"✅ Done: {out_path}")


def main():
    ok = 0
    bad = 0
    for pdf_path in PDF_PATHS:
        if not os.path.exists(pdf_path):
            print("❌ Missing PDF:", pdf_path)
            bad += 1
            continue
        out_path = make_output_path(pdf_path)
        try:
            run_one(pdf_path, out_path)
            ok += 1
        except Exception as e:
            print(f"❌ ERROR for {pdf_path}: {type(e).__name__}: {e}")
            bad += 1
    print(f"Batch complete. Success={ok} Failed={bad}")


if __name__ == "__main__":
    # Optional single-file CLI:
    #   python script.py "input.pdf" "output.xlsx"
    if len(sys.argv) >= 3:
        run_one(sys.argv[1], sys.argv[2])
    else:
        main()
