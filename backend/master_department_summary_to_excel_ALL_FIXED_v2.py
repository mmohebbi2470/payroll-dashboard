#!/usr/bin/env python3
"""
MASTER RUNNER v3 — fixes "too many lines" in Earnings by force-merging split-page rows.

Key fixes:
✅ Earnings: force merge rows by (Department, Check Date, Date) inside MASTER (even if child file changes)
✅ Optional: exclude "**** ALL ORGANIZATIONAL UNITS" from Earnings (and other tables if you want)

Run:
  python master_department_summary_to_excel_ALL_v3.py
"""

from __future__ import annotations

import os
import importlib.util
from typing import Dict, Optional

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill


# ============================================================
# MASTER CONFIG (ONLY SET THINGS HERE)
# ============================================================
PDF_PATHS = [
    # "Example Summary.pdf"
]

OUTPUT_DIR = None  # e.g. r"C:\Users\vkumawat\Downloads\Outputs"

# INCLUDE/EXCLUDE summary rows (GLOBAL)
INCLUDE_MTD = False
INCLUDE_QTD = False
INCLUDE_YTD = False

# Optional filters
EXCLUDE_ALL_ORG_UNITS = True  # drops "**** ALL ORGANIZATIONAL UNITS" rows

CREATE_SHEETS_FOR_EMPTY = True


# ============================================================
# Extractor scripts (same folder as this master file)
# ============================================================
BASE_DIR = os.path.dirname(__file__)
SCRIPT_EARNINGS = os.path.join(BASE_DIR, "new_Earnings_Same_FIXED_v2.py")
SCRIPT_DEDUCTIONS = os.path.join(BASE_DIR, "employee_deductions_universal_v22_FIXED_v2.py")
SCRIPT_WITHHOLD_LIAB = os.path.join(BASE_DIR, "withholdings_employerliab_v4d_nocalc.py")
SCRIPT_REIMB_OTHER = os.path.join(BASE_DIR, "reimbursements_otheritems_universal_FINAL_v3.py")


# ============================================================
# Helpers
# ============================================================
def _load_module(path: str, name_hint: str):
    import sys
    if not os.path.exists(path):
        raise FileNotFoundError(f"Missing required script: {path}")
    mod_name = f"_mod_{name_hint}_{abs(hash(os.path.abspath(path))) % (10**9)}"
    spec = importlib.util.spec_from_file_location(mod_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Could not import module from: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = module
    spec.loader.exec_module(module)  # type: ignore[attr-defined]
    return module


def _safe_sheet_name(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    name = pd.Series([name]).str.replace(bad, " ", regex=True).iloc[0].strip()
    return (name[:31] or "Sheet")


def _write_df_to_ws(ws, df: Optional[pd.DataFrame]):
    if df is None or df.empty:
        ws.append(["No data found"])
        return

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # numeric format
    for col in ws.iter_cols(min_row=2, min_col=1, max_col=ws.max_column):
        header = str(col[0].value or "").strip().upper()
        if header in {"CHECK DATE", "DATE", "DEPARTMENT", "OCCUR", "STATES"}:
            continue
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"


def _out_path_for_pdf(pdf_path: str) -> str:
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    out_name = f"{base}_ALL_Tables.xlsx"
    if OUTPUT_DIR:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        return os.path.join(OUTPUT_DIR, out_name)
    return os.path.join(os.path.dirname(pdf_path) or ".", out_name)


def normalize_pdf_path(p: str) -> Optional[str]:
    p = (p or "").strip().strip('"').strip("'")
    if not p:
        return None
    if os.path.exists(p):
        return p
    if not p.lower().endswith(".pdf") and os.path.exists(p + ".pdf"):
        return p + ".pdf"
    p3 = os.path.join(BASE_DIR, p)
    if os.path.exists(p3):
        return p3
    if not p.lower().endswith(".pdf"):
        p4 = os.path.join(BASE_DIR, p + ".pdf")
        if os.path.exists(p4):
            return p4
    return None


# ============================================================
# Build DataFrames
# ============================================================
def build_earnings_df(earn_mod, pdf_path: str) -> pd.DataFrame:
    """
    Why you were seeing "too many lines":
      Earnings is printed across many pages/column-bands.
      Your parser returns multiple partial rows for the SAME (Department, Check Date, Date).
      This master version FORCE-MERGES those duplicates (sums numeric parts).
    """
    df = earn_mod.parse_earnings(pdf_path)
    if df is None or df.empty:
        return pd.DataFrame()

    # normalize keys (prevents invisible whitespace causing duplicates)
    for k in ("Department", "Check Date", "Date"):
        if k in df.columns:
            df[k] = df[k].astype(str).str.strip()

    # ✅ Force merge duplicates regardless of what the child file does
    if hasattr(earn_mod, "merge_duplicate_rows"):
        df = earn_mod.merge_duplicate_rows(df)
    else:
        # fallback: merge in master
        key_cols = ["Department", "Check Date", "Date"]
        val_cols = [c for c in df.columns if c not in key_cols]
        df[val_cols] = df[val_cols].apply(pd.to_numeric, errors="coerce")
        df = df.groupby(key_cols, as_index=False)[val_cols].sum()

    # optional: exclude "ALL ORGANIZATIONAL UNITS"
    if EXCLUDE_ALL_ORG_UNITS and "Department" in df.columns:
        df = df[~df["Department"].str.upper().str.contains("ALL ORGANIZATIONAL UNITS", na=False)].reset_index(drop=True)

    # existing post-processing from your Earnings script
    df = earn_mod.rename_total_column(df, getattr(earn_mod, 'TABLE_NAME', 'EARNINGS'))
    df = earn_mod.filter_summary_rows(df)

    key_cols = ["Department", "Check Date", "Date"]
    for col in df.columns:
        if col not in key_cols:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df = earn_mod.sort_by_date(df)
    df = earn_mod.add_actual_total_earnings(df, getattr(earn_mod, 'TABLE_NAME', 'EARNINGS'))
    df = earn_mod.add_bottom_sum_row(df)
    return df


def build_employee_deductions_df(ded_mod, pdf_path: str) -> pd.DataFrame:
    # same as v2 (returns DF instead of writing workbook)
    subtables = []
    started = False
    import pdfplumber

    with pdfplumber.open(pdf_path) as pdf:
        for p_idx, page in enumerate(pdf.pages):
            page_text = (page.extract_text() or "").upper()
            if started and ("EMPLOYEE LIABILITIES" in page_text or "EMPLOYER LIABILITIES" in page_text):
                break

            blocks = ded_mod.get_deductions_blocks(page)
            if blocks:
                started = True
                for (y0, y1) in blocks:
                    subtables.extend(ded_mod.extract_subtables_from_block(page, p_idx, y0, y1))
            elif started:
                subtables.extend(ded_mod.extract_subtables_from_block(page, p_idx, 0, page.height - 5))

    if not subtables:
        return pd.DataFrame()

    sizes = [len(t["Check Date"].tolist()) for t in subtables]
    max_rows = max(sizes)
    master_idx = next(i for i, s in enumerate(sizes) if s == max_rows)
    master_src = subtables[master_idx].copy()

    master = master_src[["Check Date"] + [c for c in master_src.columns if c not in ("Check Date") and not c.startswith("_") and c != "RowPos"]]
    master = ded_mod.add_occur(master).fillna(0.0)

    parts = [t for i, t in enumerate(subtables) if i != master_idx]
    parts.sort(key=lambda d: (int(d["_page"].iloc[0]), float(d["_table_top"].iloc[0])))

    col_order = [c for c in master.columns]

    for part in parts:
        part_clean = part.copy()
        part_clean = part_clean[["Check Date"] + [c for c in part_clean.columns if c not in ("Check Date") and not c.startswith("_") and c != "RowPos"]]
        part_clean = ded_mod.add_occur(part_clean).fillna(0.0)

        if "ALL OTHER DEDUCTIONS" in part_clean.columns and "TOTAL" in part_clean.columns and "TOTAL (2)" not in part_clean.columns:
            if "TOTAL" in master.columns:
                part_clean = part_clean.rename(columns={"TOTAL": "TOTAL (2)"})

        before_cols = list(master.columns)
        master = ded_mod.merge_on_date_occur(master, part_clean)

        newly = [c for c in master.columns if c not in before_cols]
        for c in newly:
            if c not in col_order:
                col_order.append(c)

    rest = [c for c in col_order if c not in ("Check Date", "Occur")]
    final = master[["Check Date", "Occur"] + rest].fillna(0.0)

    final = ded_mod.filter_summary_rows(final)
    final = ded_mod.sort_by_check_date_label(final)
    final = ded_mod.rename_total_cols_with_prefix(final, getattr(ded_mod, 'TABLE_NAME', 'EMPLOYEE DEDUCTIONS'))

    for c in final.columns:
        if c not in ("Check Date", "Occur"):
            final[c] = pd.to_numeric(final[c], errors="coerce").fillna(0.0)

    final = ded_mod.add_actual_total_deductions(final, getattr(ded_mod, 'TABLE_NAME', 'EMPLOYEE DEDUCTIONS'))
    final = ded_mod.add_bottom_sum_row(final)
    return final


def build_withholdings_df(wl_mod, pdf_path: str) -> pd.DataFrame:
    return wl_mod.build_table(pdf_path, "EMPLOYEE WITHHOLDINGS")


def build_employer_liabilities_df(wl_mod, pdf_path: str) -> pd.DataFrame:
    return wl_mod.build_table(pdf_path, "EMPLOYER LIABILITIES")


def build_reimbursements_df(r_mod, pdf_path: str) -> pd.DataFrame:
    return r_mod.build_table(pdf_path, "REIMBURSEMENTS & OTHER PAYMENTS")


def build_other_items_df(r_mod, pdf_path: str) -> pd.DataFrame:
    return r_mod.build_table(pdf_path, "OTHER ITEMS")


def process_one_pdf(pdf_path: str, earn_mod, ded_mod, wl_mod, r_mod) -> str:
    out_xlsx = _out_path_for_pdf(pdf_path)

    tables: Dict[str, pd.DataFrame] = {
        "Earnings": build_earnings_df(earn_mod, pdf_path),
        "Employee Deductions": build_employee_deductions_df(ded_mod, pdf_path),
        "Employee Withholdings": build_withholdings_df(wl_mod, pdf_path),
        "Employer Liabilities": build_employer_liabilities_df(wl_mod, pdf_path),
        "Reimbursements & Other Payments": build_reimbursements_df(r_mod, pdf_path),
        "Other Items": build_other_items_df(r_mod, pdf_path),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, df in tables.items():
        if (df is None or df.empty) and not CREATE_SHEETS_FOR_EMPTY:
            continue
        ws = wb.create_sheet(title=_safe_sheet_name(sheet_name))
        _write_df_to_ws(ws, df)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Data")
        ws.append(["No data found in any section."])

    wb.save(out_xlsx)
    return out_xlsx


def main():
    earn_mod = __import_module(SCRIPT_EARNINGS, "earnings")
    ded_mod = __import_module(SCRIPT_DEDUCTIONS, "deductions")
    wl_mod = __import_module(SCRIPT_WITHHOLD_LIAB, "withhold_liab")
    r_mod = __import_module(SCRIPT_REIMB_OTHER, "reimb_other")

    # push global flags into modules
    for mod in (earn_mod, ded_mod, wl_mod, r_mod):
        mod.INCLUDE_MTD = INCLUDE_MTD
        mod.INCLUDE_QTD = INCLUDE_QTD
        mod.INCLUDE_YTD = INCLUDE_YTD

    ok, bad = 0, 0
    for raw in PDF_PATHS:
        pdf_path = normalize_pdf_path(raw)
        if not pdf_path:
            print("❌ Missing PDF:", raw)
            bad += 1
            continue
        try:
            out_xlsx = process_one_pdf(pdf_path, earn_mod, ded_mod, wl_mod, r_mod)
            print("✅ Saved:", out_xlsx)
            ok += 1
        except Exception as e:
            print(f"❌ ERROR for {pdf_path}: {type(e).__name__}: {e}")
            bad += 1
    print(f"Done. Success={ok} Failed={bad}")


def __import_module(path: str, hint: str):
    return _load_module(path, hint)


if __name__ == "__main__":
    main()