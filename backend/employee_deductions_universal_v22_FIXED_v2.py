# employee_deductions_universal_v15_batch.py
# ✅ v14 logic + requested changes + MULTI-PDF batch

import os
import sys
import re
from datetime import datetime

import pdfplumber
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# =========================
# CONFIG (hardcode here)
# =========================
TABLE_NAME = "EMPLOYEE DEDUCTIONS"
PDF_PATHS = [
    r"PG January 2026 Department Summary.pdf",


]

# If None -> output next to each PDF
OUTPUT_DIR = None  # e.g. r"C:\Users\vkumawat\Downloads\Outputs"

# =========================
# SUMMARY ROW CONTROL
# =========================
INCLUDE_MTD = False
INCLUDE_QTD = False
INCLUDE_YTD = False
# # MULTI PDF INPUTS (hardcode all PDFs here)
# PDF_PATHS = [
#     r"MC Department summary 1.pdf",
#     r"MC Department summary 2.pdf",
#     r"MC Department summary 3.pdf",
#     r"PG Department summary 1.pdf",
#     r"PG Department summary 2.pdf",
#     r"PG Department summary 3.pdf",

# ]
# # If None -> output next to each PDF
# OUTPUT_DIR = None  # e.g. r"C:\Users\vkumawat\Downloads\Outputs"

# # INCLUDE/EXCLUDE summary rows
# INCLUDE_MTD = False
# INCLUDE_QTD = False
# INCLUDE_YTD = False

# =========================
# Regex
# =========================
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2}$")
ROWLABEL_RE = re.compile(r"^(MTD|QTD|YTD)\s*\(.*\)$|^(MTD|QTD|YTD)\(.*\)$", re.I)
NUM_RE = re.compile(r"^-?\d[\d,]*\.\d{2}$")
X_CLUSTER_THRESHOLD = 26


def get_check_date(page):
    t = page.extract_text(layout=True) or ""
    m = re.search(r"Check\s*Date\s+(\d{2}/\d{2}/\d{2})", t, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"CheckDate\s+(\d{2}/\d{2}/\d{2})", t, flags=re.IGNORECASE)
    return m.group(1) if m else None


def to_float(s: str) -> float:
    try:
        return float(str(s).replace(",", ""))
    except Exception:
        return 0.0


def group_lines(words, y_tol=2.0):
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines, cur, cur_y = [], [], None
    for w in words:
        y = w["top"]
        if cur_y is None or abs(y - cur_y) <= y_tol:
            cur.append(w)
            cur_y = y if cur_y is None else (cur_y + y) / 2
        else:
            lines.append(sorted(cur, key=lambda x: x["x0"]))
            cur, cur_y = [w], y
    if cur:
        lines.append(sorted(cur, key=lambda x: x["x0"]))
    return lines


def normalize_header_token(t: str) -> str:
    m = {
        "401K": "401 K",
        "401KLOAN": "401K LOAN",
        "401KEECATCH": "401K EE CATCH UP",
        "AFLACPOST": "AFLAC POST TAX",
        "AFLACSTD": "AFLAC STD POST TAX",
        "FSAMEDICAL": "FSA MEDICAL",
        "UNIONDUE": "UNION DUE",
        "SHOPCHARGES": "SHOP CHARGES",
        "ROTH401KEE": "ROTH 401K EE",
        "ALLOTHER": "ALL OTHER DEDUCTIONS",
        "ALLOTHERWITHHOLDINGS": "ALL OTHER DEDUCTIONS",
    }
    return m.get(t.strip().upper(), t.strip())


def is_valid_header_token(t: str) -> bool:
    if not t:
        return False
    tu = t.strip().upper()
    if tu in ("CHECK", "DATE", "MTD", "QTD", "YTD"):
        return False
    if DATE_RE.match(t) or NUM_RE.match(t):
        return False
    if len(tu) == 1:
        return False
    if "CONT" in tu:
        return False
    if "EMPLOYEEDEDUCTIONS" in tu.replace(" ", ""):
        return False
    if tu in ("PAGE", "DEPARTMENT", "ORGANIZATION", "UNIT", "SUMMARY"):
        return False
    return re.search(r"[A-Z]", tu) is not None


def make_unique_in_order(names):
    counts = {}
    out = []
    for n in names:
        if n not in counts:
            counts[n] = 1
            out.append(n)
        else:
            counts[n] += 1
            out.append(f"{n} ({counts[n]})" if n.upper() != "TOTAL" else f"TOTAL ({counts[n]})")
    return out


def detect_columns_by_x_center(header_words, x_thresh=X_CLUSTER_THRESHOLD):
    cleaned = []
    for w in header_words:
        t = (w.get("text") or "").strip()
        if not t:
            continue

        is_suffix_num = bool(re.fullmatch(r"\d{1,2}", t))

        if not is_suffix_num and not is_valid_header_token(t):
            continue

        if (not is_suffix_num) and (DATE_RE.match(t) or NUM_RE.match(t)):
            continue

        x_center = (w["x0"] + w["x1"]) / 2
        cleaned.append({**w, "x_center": x_center, "text": normalize_header_token(t)})

    if not cleaned:
        return []

    cleaned = sorted(cleaned, key=lambda x: x["x_center"])
    clusters = []
    cur = [cleaned[0]]
    for w in cleaned[1:]:
        if abs(w["x_center"] - cur[-1]["x_center"]) <= x_thresh:
            cur.append(w)
        else:
            clusters.append(cur)
            cur = [w]
    clusters.append(cur)

    cols = []
    for c in clusters:
        c_sorted = sorted(c, key=lambda x: (x["top"], x["x0"]))
        title_tokens = [x["text"] for x in c_sorted]
        title = " ".join(title_tokens)
        title = " ".join(title.split()).strip()
        title = re.sub(r"\b(ALL OTHER DEDUCTIONS)\s+DEDUCTIONS\b", r"\1", title, flags=re.I)

        mid = sum(x["x_center"] for x in c) / len(c)
        tu = title.upper()
        if "PAYACTIV" in tu and "READYCHEX" in tu:
            cols.append(("PAYACTIV", mid - 25))
            cols.append(("READYCHEX FEE", mid + 25))
        else:
            cols.append((title, mid))
    return cols


def find_section_headers(page):
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    lines = group_lines(words)
    headers = []
    for ln in lines:
        text = " ".join([w["text"] for w in ln]).strip()
        up = text.upper().replace(" ", "")
        if "EMPLOYEEDEDUCTIONS" in up or ("EMPLOYEE" in text.upper() and "DEDUCTIONS" in text.upper()):
            headers.append(("EMPLOYEE DEDUCTIONS", ln[0]["top"]))
        if "LIABILITIES" in up and "EMPLOYEE" in up:
            headers.append(("EMPLOYEE LIABILITIES", ln[0]["top"]))
        if "LIABILITIES" in up and "EMPLOYER" in up:
            headers.append(("EMPLOYER LIABILITIES", ln[0]["top"]))
    headers = sorted(headers, key=lambda x: x[1])
    cleaned = []
    for n, y in headers:
        if not cleaned or abs(y - cleaned[-1][1]) > 5 or n != cleaned[-1][0]:
            cleaned.append((n, y))
    return cleaned


def get_deductions_blocks(page):
    hs = find_section_headers(page)
    blocks = []
    for i, (name, y) in enumerate(hs):
        if name != "EMPLOYEE DEDUCTIONS":
            continue
        y0 = y - 2
        y1 = page.height - 5
        for j in range(i + 1, len(hs)):
            if hs[j][1] > y:
                y1 = hs[j][1] - 2
                break
        blocks.append((y0, y1))
    return blocks


def extract_subtables_from_block(page, page_idx, y0, y1):
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    seg = [w for w in words if y0 <= w["top"] <= y1]
    if not seg:
        return []

    early = " ".join([w["text"] for w in seg if w["top"] <= y0 + 120]).upper().replace(" ", "")
    if "LIABILITIES" in early:
        return []

    lines = group_lines(seg)

    check_idxs = []
    for i, ln in enumerate(lines):
        if ln and ln[0]["text"].strip().upper() == "CHECK":
            ok = False
            for j in range(i + 1, min(i + 12, len(lines))):
                if lines[j] and lines[j][0]["text"].strip().upper() == "DATE":
                    ok = True
                    break
            if ok:
                check_idxs.append(i)

    subtables = []
    for k, ci in enumerate(check_idxs):
        check_ln = lines[ci]
        date_ln = None
        for j in range(ci + 1, min(ci + 12, len(lines))):
            if lines[j] and lines[j][0]["text"].strip().upper() == "DATE":
                date_ln = lines[j]
                break
        if date_ln is None:
            continue

        header_top = max(y0, min(w["top"] for w in check_ln) - 160)
        header_bottom = max(w["bottom"] for w in date_ln) + 3
        header_band = [w for w in seg if header_top <= w["top"] <= header_bottom]

        cols = detect_columns_by_x_center(header_band)
        if not cols:
            continue
        col_names = make_unique_in_order([c[0] for c in cols])
        col_mids = {col_names[i]: cols[i][1] for i in range(len(cols))}

        row_y0 = header_bottom + 1
        next_ci = check_idxs[k + 1] if k + 1 < len(check_idxs) else None
        row_y1 = y1
        if next_ci is not None:
            row_y1 = min(row_y1, lines[next_ci][0]["top"] - 2)

        rows = []
        rowpos = 0
        for ln in lines:
            if not ln:
                continue
            if ln[0]["top"] < row_y0 or ln[0]["top"] > row_y1:
                continue
            first = ln[0]["text"].strip()
            fu = first.upper().replace(" ", "")

            if DATE_RE.match(first) or ROWLABEL_RE.match(first.upper()) or ROWLABEL_RE.match(fu):
                label = first
                # If the first token isn't a valid date (e.g. MTD), try page-level date as fallback
                if not DATE_RE.match(label):
                    cd_fallback = get_check_date(page)
                    if cd_fallback:
                        label = cd_fallback

                rowpos += 1
                row = {"RowPos": rowpos, "Check Date": label, "_page": page_idx}
                for c in col_mids:
                    row[c] = 0.0
                nums = [w for w in ln[1:] if NUM_RE.match(w["text"].strip())]
                for n in nums:
                    val = to_float(n["text"].strip())
                    x_mid = (n["x0"] + n["x1"]) / 2
                    closest = min(col_mids.items(), key=lambda kv: abs(x_mid - kv[1]))[0]
                    row[closest] += val
                rows.append(row)

        if not rows:
            continue

        df = pd.DataFrame(rows).fillna(0.0)
        df["_rowseq"] = "|".join(df["Check Date"].astype(str).tolist())
        df["_table_top"] = float(check_ln[0]["top"])
        subtables.append(df)

    subtables.sort(key=lambda d: float(d["_table_top"].iloc[0]))
    return subtables


def add_occur(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Occur"] = df.groupby("Check Date").cumcount() + 1
    return df


def merge_on_date_occur(master: pd.DataFrame, part: pd.DataFrame) -> pd.DataFrame:
    key_cols = ["Check Date", "Occur"]
    m = master.copy()
    p = part.copy()

    if "Occur" not in m.columns:
        m = add_occur(m)
    if "Occur" not in p.columns:
        p = add_occur(p)

    bring = [c for c in p.columns if c not in key_cols and c != "RowPos" and not c.startswith("_")]
    merged = pd.merge(m, p[key_cols + bring], on=key_cols, how="left", suffixes=("", "__dup")).fillna(0.0)

    for c in list(merged.columns):
        if c.endswith("__dup"):
            base = c[:-5].strip()
            if base in merged.columns:
                merged[base] = pd.to_numeric(merged[base], errors="coerce").fillna(0.0) + pd.to_numeric(
                    merged[c], errors="coerce"
                ).fillna(0.0)
                merged.drop(columns=[c], inplace=True)
                continue

            base_u = base.upper()
            k = 2
            new = f"{base} ({k})" if base_u != "TOTAL" else f"TOTAL ({k})"
            while new in merged.columns:
                k += 1
                new = f"{base} ({k})" if base_u != "TOTAL" else f"TOTAL ({k})"
            merged.rename(columns={c: new}, inplace=True)

    return merged


# =========================
# NEW: filtering/sorting/totals
# =========================
def filter_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    cd = out["Check Date"].astype(str).str.upper()
    mask = pd.Series(True, index=out.index)
    if not INCLUDE_MTD:
        mask &= ~cd.str.startswith("MTD", na=False)
    if not INCLUDE_QTD:
        mask &= ~cd.str.startswith("QTD", na=False)
    if not INCLUDE_YTD:
        mask &= ~cd.str.startswith("YTD", na=False)
    out = out.loc[mask].reset_index(drop=True)
    return out


def sort_key_label(label: str):
    s = (label or "").strip().upper()
    if DATE_RE.match(s):
        try:
            return (0, datetime.strptime(s, "%m/%d/%y"))
        except Exception:
            return (0, datetime.max)
    if s.startswith("MTD"):
        return (1, s)
    if s.startswith("QTD"):
        return (2, s)
    if s.startswith("YTD"):
        return (3, s)
    return (4, s)


def sort_by_check_date_label(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["_k"] = out["Check Date"].apply(sort_key_label)
    out = out.sort_values(["_k", "Occur"], kind="mergesort").drop(columns=["_k"]).reset_index(drop=True)
    return out


def rename_total_cols_with_prefix(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    """
    TOTAL -> "<table> Total"
    TOTAL (2) -> "<table> Total (2)"
    etc.
    """
    out = df.copy()
    ren = {}
    for c in out.columns:
        cu = str(c).strip().upper()
        if cu == "TOTAL" or cu.startswith("TOTAL ("):
            suffix = "" if cu == "TOTAL" else c[len("TOTAL") :]  # keeps " (2)" etc
            ren[c] = f"{table_name} Total{suffix}"
    return out.rename(columns=ren)


def add_actual_total_deductions(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    """
    "Actual Total Deductions" = sum of ALL numeric deduction columns EXCEPT the Total columns.
    """
    key_cols = {"Check Date", "Occur"}
    total_prefix = f"{table_name.strip().upper()} TOTAL"

    def is_total_col(col: str) -> bool:
        up = str(col).strip().upper()
        return up == total_prefix or up.startswith(total_prefix + " (")

    numeric_cols = [c for c in df.columns if c not in key_cols and not is_total_col(c)]

    out = df.copy()
    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    out["Actual Total Deductions"] = out[numeric_cols].sum(axis=1)
    return out


def collapse_same_check_date(df: pd.DataFrame) -> pd.DataFrame:
    """
    If a page continuation causes the SAME check date to appear multiple times,
    collapse into ONE row per check date by summing numeric columns.

    Keeps MTD/QTD/YTD rows separate (those are not real dates).
    """
    out = df.copy()
    if "Check Date" not in out.columns:
        return out

    s = out["Check Date"].astype(str).str.strip()
    is_date = s.apply(lambda x: bool(DATE_RE.match(x)))

    date_rows = out[is_date].copy()
    other_rows = out[~is_date].copy()

    # numeric columns = everything except identifiers
    id_cols = [c for c in ("Check Date", "Occur") if c in out.columns]
    num_cols = [c for c in out.columns if c not in id_cols]

    for c in num_cols:
        date_rows[c] = pd.to_numeric(date_rows[c], errors="coerce").fillna(0.0)

    collapsed = date_rows.groupby("Check Date", as_index=False)[num_cols].sum()
    if "Occur" in out.columns:
        collapsed.insert(1, "Occur", 1)

    combined = pd.concat([collapsed, other_rows], ignore_index=True)

    if "Occur" in combined.columns:
        combined["Occur"] = pd.to_numeric(combined["Occur"], errors="coerce").fillna(1).astype(int)

    return combined



def add_bottom_sum_row(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    sum_row = {c: "" for c in out.columns}
    sum_row["Check Date"] = "TOTAL"
    if "Occur" in out.columns:
        sum_row["Occur"] = ""

    for c in out.columns:
        if c in ("Check Date", "Occur"):
            continue
        sum_row[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).sum()

    return pd.concat([out, pd.DataFrame([sum_row])], ignore_index=True)


# =========================
# Core run for ONE PDF
# =========================
def run_one(pdf_path: str, out_path: str):
    subtables = []
    started = False

    with pdfplumber.open(pdf_path) as pdf:
        for p_idx, page in enumerate(pdf.pages):
            page_text = (page.extract_text() or "").upper()
            if started and ("EMPLOYEE LIABILITIES" in page_text or "EMPLOYER LIABILITIES" in page_text):
                break

            blocks = get_deductions_blocks(page)

            if blocks:
                started = True
                for (y0, y1) in blocks:
                    subtables.extend(extract_subtables_from_block(page, p_idx, y0, y1))
            elif started:
                subtables.extend(extract_subtables_from_block(page, p_idx, 0, page.height - 5))

    if not subtables:
        print("No Employee Deductions found:", pdf_path)
        return

    sizes = [len(t["Check Date"].tolist()) for t in subtables]
    max_rows = max(sizes)
    master_idx = next(i for i, s in enumerate(sizes) if s == max_rows)
    master_src = subtables[master_idx].copy()

    master = master_src[["Check Date"] + [c for c in master_src.columns if c not in ("Check Date") and not c.startswith("_") and c != "RowPos"]]
    master = add_occur(master).fillna(0.0)

    parts = [t for i, t in enumerate(subtables) if i != master_idx]
    parts.sort(key=lambda d: (int(d["_page"].iloc[0]), float(d["_table_top"].iloc[0])))

    col_order = [c for c in master.columns]

    for part in parts:
        part_clean = part.copy()
        part_clean = part_clean[["Check Date"] + [c for c in part_clean.columns if c not in ("Check Date") and not c.startswith("_") and c != "RowPos"]]
        part_clean = add_occur(part_clean).fillna(0.0)

        if "ALL OTHER DEDUCTIONS" in part_clean.columns and "TOTAL" in part_clean.columns and "TOTAL (2)" not in part_clean.columns:
            if "TOTAL" in master.columns:
                part_clean = part_clean.rename(columns={"TOTAL": "TOTAL (2)"})

        before_cols = list(master.columns)
        master = merge_on_date_occur(master, part_clean)

        newly = [c for c in master.columns if c not in before_cols]
        for c in newly:
            if c not in col_order:
                col_order.append(c)

    rest = [c for c in col_order if c not in ("Check Date", "Occur")]
    final = master[["Check Date", "Occur"] + rest].fillna(0.0)

    # ===== apply requested transformations =====
    final = filter_summary_rows(final)
    final = sort_by_check_date_label(final)
    final = rename_total_cols_with_prefix(final, TABLE_NAME)

    # numeric formatting / fill
    for c in final.columns:
        if c not in ("Check Date", "Occur"):
            final[c] = pd.to_numeric(final[c], errors="coerce").fillna(0.0)

    final = add_actual_total_deductions(final, TABLE_NAME)
    final = collapse_same_check_date(final)
    final = add_bottom_sum_row(final)

    # ===== write excel =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Deductions"

    for row in dataframe_to_rows(final, index=False, header=True):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # number format for numeric columns (starting col 3 because col1=Check Date, col2=Occur)
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=ws.max_column):
        for cell in col:
            cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"✅ Done: {out_path}")


def make_output_path(pdf_path: str) -> str:
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    out_name = f"{base}_Employee_Deductions.xlsx"
    if OUTPUT_DIR:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        return os.path.join(OUTPUT_DIR, out_name)
    return os.path.join(os.path.dirname(pdf_path) or ".", out_name)


def main():
    ok = 0
    bad = 0
    for pdf_path in PDF_PATHS:
        if not os.path.exists(pdf_path):
            print("❌ Missing PDF:", pdf_path)
            bad += 1
            continue
        out_path = make_output_path(pdf_path)
        try:
            run_one(pdf_path, out_path)
            ok += 1
        except Exception as e:
            print(f"❌ ERROR for {pdf_path}: {type(e).__name__}: {e}")
            bad += 1
    print(f"Batch complete. Success={ok} Failed={bad}")


if __name__ == "__main__":
    # Optional CLI: python script.py input.pdf output.xlsx (runs single file)
    if len(sys.argv) >= 3:
        run_one(sys.argv[1], sys.argv[2])
    else:
        main()