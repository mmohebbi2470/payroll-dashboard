#!/usr/bin/env python3
"""
AUTO FILL VIJAY PAYROLL v8 (FINAL PRODUCTION VERSION)

What this fixes (per your MC output + template behavior):
✅ Fills ALL rows that have formulas (including the bottom "Employer Taxes" row).
✅ Skips ONLY special summary rows:
   - any row whose Column B contains "TOTAL PAYROLL" (case-insensitive)
   - the row whose Column B contains "NET OF PAYROLL" (case-insensitive)
✅ Computes and writes:
   - Row totals: Column F (Cr) and Column G (Db)
✅ Writes summary values:
   - TOTAL PAYROLL row(s): sums of ALL regular rows (excludes TOTAL PAYROLL + NET OF PAYROLL rows)
   - NET OF PAYROLL row: (Total Cr - Total Db)
✅ Keeps your v5 mapping behavior:
   - If token missing => 0 (NO roll-up to "ALL OTHER")
   - Token aliases across PG/MC supported (any / sum)

Usage:
  1) Set CONFIG paths + TARGET_PAY_DATE
  2) python auto_fill_vijay_payroll_from_all_tables_v8_final.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union, Literal

import pandas as pd
import openpyxl
import sys
import os


# =========================
# CONFIG (HARD CODE HERE)
# =========================
TARGET_PAY_DATE = "01/16/26"
ALL_TABLES_XLSX = "source_data.xlsx" 
PAYROLL_TEMPLATE_XLSX = "Vijay Payroll.xlsx"

PAYROLL_SHEET_NAME = None  # None = active sheet

COL_FORMULA_DEBT = "D"   # formula that fills Column F (Cr)
COL_FORMULA_CREDIT = "E" # formula that fills Column G (Db)
COL_OUT_DEBT = "F"       # Cr
COL_OUT_CREDIT = "G"     # Db

DEDUP_TOKENS_IN_FORMULA = True
BLANK_ROW_STOP_STREAK = 50  # keep high; template may have blanks between sections


# =========================
# TOKEN -> COLUMN MAPPING
# =========================
MappingVal = Union[str, Dict[Literal["any", "sum"], List[str]]]

ALIASES: Dict[str, MappingVal] = {
    # --- Common spelling fixes / header variants ---
    "ALL OTHER EARBINGS": {"any": ["ALL OTHER"]},
    "DONATED VAVATION REC": {"any": ["DONATED VACATIONREC", "DONATED VACATION REC"]},

    # --- Earnings: PG vs MC header differences ---
    "TRAINING OT": {"any": ["TRAININGOT", "TRAININGOT HOURS", "TRAINING OT", "TRAINING OT HOURS"]},
    "TRAINING OT HOURS": {"any": ["TRAININGOT HOURS", "TRAININGOT", "TRAINING OT HOURS", "TRAINING OT"]},

    "RE-TRAINING": {"any": ["RE-TRAINING"]},
    "STAND-BY": {"any": ["STAND-BY"]},

    "COVID CHILDCARE": {"any": ["COVIDCHILDCARE", "COVID CHILDCARE"]},
    "COVID SICK-EE": {"any": ["COVIDSICK-EE", "COVID SICK-EE"]},
    "COVID SICK-FAM": {"any": ["COVIDSICK-FAM", "COVID SICK-FAM"]},

    # BONUS breakdown exists in MC; PG sometimes has only BONUS
    "BONUS-ATTENDANCE": {"any": ["Bonus-Attendance", "BONUS-ATTENDANCE", "BONUS ATTENDANCE"]},
    "BONUS-DISCRETIONARY": {"any": ["Bonus-Discretionary", "BONUS-DISCRETIONARY", "BONUS DISCRETIONARY"]},
    "BONUS-TRAINER": {"any": ["Bonus-Trainer", "BONUS-TRAINER", "BONUS TRAINER"]},
    "BONUS": {"any": ["BONUS"]},

    "HOLIDAY REG": {"any": ["HOLIDAYREG", "HOLIDAY REG"]},

    # --- Deductions/Withholdings/Liabilities totals ---
    # (Add multiple misspellings you used in formulas)
    "EMPLOYEE WITHHOLDING TOTAL": {"any": ["EMPLOYEE WITHHOLDINGS Total", "EMPLOYEE WITHHOLDINGS TOTAL"]},
    "EMPLOYEE WITHHOLDINGS TOTAL": {"any": ["EMPLOYEE WITHHOLDINGS Total", "EMPLOYEE WITHHOLDINGS TOTAL"]},

    "EMPLOYER LIABILITIES TOTAL": {"any": ["EMPLOYER LIABILITIES Total", "EMPLOYER LIABILITIES TOTAL"]},
    "EMPLOYER LIABILITY TOTAL": {"any": ["EMPLOYER LIABILITIES Total", "EMPLOYER LIABILITIES TOTAL"]},
    "EMPOYER LIABILITIES TOTAL": {"any": ["EMPLOYER LIABILITIES Total", "EMPLOYER LIABILITIES TOTAL"]},
    "EMPOYEE LIABILTY TOTAL": {"any": ["EMPLOYER LIABILITIES Total", "EMPLOYER LIABILITIES TOTAL"]},  # your typo
    "EMPLOYER LIABAILTIES TOTAL": {"any": ["EMPLOYER LIABILITIES Total", "EMPLOYER LIABILITIES TOTAL"]},  # old typo
    "EMPLOYER TOTAL LIABILITIES TOTAL": {"any": ["EMPLOYER LIABILITIES Total", "EMPLOYER LIABILITIES TOTAL"]},

    # Your new requested token names
    "TOTAL PAYROLL DEBTS": {"any": ["TOTAL PAYROLL DEBTS"]},  # if exists as its own column; otherwise stays 0

    "CELL PHONE REIMBE": {"any": ["CELL PHONE REIMB", "CELL PHONE REIMBE"]},
    "1100 MISC PAY": {"any": ["1099 MISC COMP", "1100 MISC PAY"]},

    # SUM example
    "401K employee": {"sum": ["401 K", "401K EE CATCH UP UP", "401K EE CATCHUP", "ROTH 401K EE", "ROTH 401K EE CATCHUP"]},
}


# =========================
# Normalization helpers
# =========================
def norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def norm_key(s: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (s or "").upper())


def safe_date_str(val) -> str:
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val).strip().split(" ")[0]
    try:
         dt = pd.to_datetime(val_str)
         return dt.strftime("%m/%d/%y").lstrip("0").replace("/0", "/")
    except:
         parts = val_str.split("/")
         if len(parts) == 3:
             m, d, y = parts
             y = y[-2:] if len(y) == 4 else y
             return f"{int(m)}/{int(d)}/{y}"
         return val_str


def to_float(x) -> float:
    try:
        if x is None:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().replace(",", "")
        if s == "" or s.lower() == "nan":
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def parse_formula(expr: str) -> List[str]:
    if not expr:
        return []
    parts = [norm_space(p) for p in str(expr).split("+")]
    parts = [p for p in parts if p]

    if not DEDUP_TOKENS_IN_FORMULA:
        return parts

    seen = set()
    out = []
    for p in parts:
        k = norm_key(p)
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out


def resolve_mapping(token: str) -> Tuple[str, List[str]]:
    """
    Returns (mode, candidates)
      mode = "any" or "sum"
      candidates = list of column header candidates to try
    """
    t = norm_space(token)
    if not t:
        return "any", []

    # direct key match
    if t in ALIASES:
        v = ALIASES[t]
    else:
        # normalized key match (handles case/space differences)
        nk = norm_key(t)
        v = None
        for k, vv in ALIASES.items():
            if norm_key(k) == nk:
                v = vv
                break
        if v is None:
            return "any", [t]

    if isinstance(v, str):
        return "any", [norm_space(v)]
    if isinstance(v, dict):
        if "sum" in v:
            return "sum", [norm_space(x) for x in v["sum"] if norm_space(x)]
        return "any", [norm_space(x) for x in v.get("any", []) if norm_space(x)]
    return "any", [t]


# =========================
# Department matching
# =========================
CONT_RE = re.compile(r"\(CONT\.?\)|CONT\.?$", re.I)

def dept_keys(s: str) -> List[str]:
    s = (s or "").upper()
    s = s.replace("*", " ")
    s = CONT_RE.sub("", s)
    s = norm_space(s)
    k = norm_key(s)
    keys = []
    if k:
        keys.append(k)
        if k.endswith("S") and len(k) > 3:
            keys.append(k[:-1])
    return keys


def split_payroll_departments(payroll_dept: str) -> List[str]:
    s = norm_space(payroll_dept or "")
    if not s:
        return []
    parts = re.split(r"\s*(?:\+|&|,|/|;|\band\b)\s*", s, flags=re.I)
    return [norm_space(p) for p in parts if norm_space(p)]


def dept_matches(extracted_dept: str, payroll_dept_part: str) -> bool:
    if not payroll_dept_part:
        return True
    ek = dept_keys(extracted_dept)
    pk = dept_keys(payroll_dept_part)
    if not ek or not pk:
        return False

    for e in ek:
        for p in pk:
            if p in e or e in p:
                return True

    m = re.search(r"\b(\d{2,5})\b", payroll_dept_part)
    if m:
        code = m.group(1)
        for e in ek:
            if code in e:
                return True
    return False


# =========================
# Extracted workbook model
# =========================
@dataclass
class SheetInfo:
    name: str
    df: pd.DataFrame
    date_col: str
    dept_col: Optional[str]


def load_all_tables(path: str) -> Dict[str, SheetInfo]:
    xls = pd.ExcelFile(path, engine="openpyxl")
    infos: Dict[str, SheetInfo] = {}
    for sh in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sh, engine="openpyxl")
        df.columns = [norm_space(str(c)) for c in df.columns]

        if "Date" in df.columns:
            date_col = "Date"
        elif "Check Date" in df.columns:
            date_col = "Check Date"
        else:
            date_col = ""

        dept_col = "Department" if "Department" in df.columns else None

        if dept_col:
            df[dept_col] = df[dept_col].astype(str).map(norm_space)
        if date_col:
            df[date_col] = df[date_col].apply(safe_date_str)

        infos[sh] = SheetInfo(name=sh, df=df, date_col=date_col, dept_col=dept_col)
    return infos


def build_column_index(infos: Dict[str, SheetInfo]) -> Dict[str, List[Tuple[str, str]]]:
    idx: Dict[str, List[Tuple[str, str]]] = {}
    for sh, info in infos.items():
        for c in info.df.columns:
            k = norm_key(c)
            if not k:
                continue
            idx.setdefault(k, []).append((sh, c))
    return idx


def filter_rows(info: SheetInfo, payroll_dept: str, target_date: str) -> pd.DataFrame:
    df = info.df
    if info.date_col:
        df = df[df[info.date_col] == safe_date_str(target_date)]
    if info.dept_col:
        parts = split_payroll_departments(payroll_dept)
        if parts:
            col = info.dept_col
            mask_any = df[col].astype(str).apply(lambda x: any(dept_matches(x, p) for p in parts))
            df = df[mask_any]
    return df


def token_value_single_column(
    col_name: str,
    payroll_dept: str,
    target_date: str,
    infos: Dict[str, SheetInfo],
    col_index: Dict[str, List[Tuple[str, str]]],
) -> Tuple[float, Optional[str], Optional[str]]:
    k = norm_key(col_name)
    if not k or k not in col_index:
        return 0.0, None, None

    preferred = [
        "Earnings",
        "Employee Deductions",
        "Employee Withholdings",
        "Employer Liabilities",
        "Reimbursements & Other Payments",
        "Other Items",
    ]
    candidates = col_index[k]

    def sort_key(item):
        sh, _ = item
        return (preferred.index(sh) if sh in preferred else 999, sh)

    for sh, actual_col in sorted(candidates, key=sort_key):
        info = infos[sh]
        if not info.date_col:
            continue
        sub = filter_rows(info, payroll_dept, target_date)
        if sub.empty:
            continue
        val = sub[actual_col].map(to_float).sum()
        return float(val), sh, actual_col

    # column exists but dept/date missing -> 0 with reference
    sh0, col0 = sorted(candidates, key=sort_key)[0]
    return 0.0, sh0, col0


def token_value(
    token: str,
    payroll_dept: str,
    target_date: str,
    infos: Dict[str, SheetInfo],
    col_index: Dict[str, List[Tuple[str, str]]],
) -> Tuple[float, Optional[str], Optional[str]]:
    mode, candidates = resolve_mapping(token)
    if not candidates:
        return 0.0, None, None

    if mode == "sum":
        total = 0.0
        used = []
        first_sh = None
        first_col = None
        for c in candidates:
            v, sh, col = token_value_single_column(c, payroll_dept, target_date, infos, col_index)
            total += v
            if sh is not None:
                used.append(c)
                if first_sh is None:
                    first_sh, first_col = sh, col
        if first_sh is None:
            return 0.0, None, None
        return float(total), first_sh, f"{first_col}  (sum: {', '.join(used)})"

    # mode == "any"
    for c in candidates:
        v, sh, col = token_value_single_column(c, payroll_dept, target_date, infos, col_index)
        if sh is not None:
            return float(v), sh, col

    return 0.0, None, None


# =========================
# Payroll fill
# =========================
def get_payroll_sheet(wb: openpyxl.Workbook):
    if PAYROLL_SHEET_NAME and PAYROLL_SHEET_NAME in wb.sheetnames:
        return wb[PAYROLL_SHEET_NAME]
    return wb.active


def row_is_blank(ws, r: int) -> bool:
    a = ws[f"A{r}"].value
    b = ws[f"B{r}"].value
    c = ws[f"C{r}"].value
    d = ws[f"{COL_FORMULA_DEBT}{r}"].value
    e = ws[f"{COL_FORMULA_CREDIT}{r}"].value
    return all((v is None or str(v).strip() == "") for v in (a, b, c, d, e))


def is_total_payroll_row(ws, r: int) -> bool:
    a = str(ws[f"A{r}"].value or "").strip().lower()
    b = str(ws[f"B{r}"].value or "").strip().lower()
    return "total payroll" in a or "total payroll" in b


def is_net_row(ws, r: int) -> bool:
    a = str(ws[f"A{r}"].value or "").strip().lower()
    b = str(ws[f"B{r}"].value or "").strip().lower()
    return "net of payroll" in a or "net of payroll" in b


def main():
    global TARGET_PAY_DATE, OUTPUT_XLSX

    infos = load_all_tables(ALL_TABLES_XLSX)
    col_index = build_column_index(infos)

    wb = openpyxl.load_workbook(PAYROLL_TEMPLATE_XLSX)
    ws = get_payroll_sheet(wb)

    debug_rows = []
    missing_tokens = set()

    max_r = ws.max_row
    blank_streak = 0

    # Fill all rows EXCEPT summary rows (TOTAL PAYROLL / NET OF PAYROLL)
    for r in range(2, max_r + 1):
        if row_is_blank(ws, r):
            blank_streak += 1
            if blank_streak >= BLANK_ROW_STOP_STREAK:
                break
            continue
        blank_streak = 0

        if is_total_payroll_row(ws, r) or is_net_row(ws, r):
            continue

        payroll_dept = norm_space(str(ws[f"C{r}"].value or ""))

        debt_expr = norm_space(str(ws[f"{COL_FORMULA_DEBT}{r}"].value or ""))
        cred_expr = norm_space(str(ws[f"{COL_FORMULA_CREDIT}{r}"].value or ""))

        debt_total = 0.0
        for tok in parse_formula(debt_expr):
            v, sh, col = token_value(tok, payroll_dept, TARGET_PAY_DATE, infos, col_index)
            debt_total += v
            debug_rows.append([r, "DEBT->Cr", payroll_dept, tok, v, sh, col, TARGET_PAY_DATE])
            if sh is None:
                missing_tokens.add(tok)

        cred_total = 0.0
        for tok in parse_formula(cred_expr):
            v, sh, col = token_value(tok, payroll_dept, TARGET_PAY_DATE, infos, col_index)
            cred_total += v
            debug_rows.append([r, "CREDIT->Db", payroll_dept, tok, v, sh, col, TARGET_PAY_DATE])
            if sh is None:
                missing_tokens.add(tok)

        ws[f"{COL_OUT_DEBT}{r}"].value = float(round(debt_total, 2))
        ws[f"{COL_OUT_CREDIT}{r}"].value = float(round(cred_total, 2))

    # Compute totals (excluding TOTAL PAYROLL rows + NET row)
    total_cr = 0.0
    total_db = 0.0
    for r in range(2, max_r + 1):
        if row_is_blank(ws, r):
            continue
        if is_total_payroll_row(ws, r) or is_net_row(ws, r):
            continue
        total_cr += to_float(ws[f"{COL_OUT_DEBT}{r}"].value)
        total_db += to_float(ws[f"{COL_OUT_CREDIT}{r}"].value)

    # Write totals into ALL "TOTAL PAYROLL" rows found
    # NOTE: If we want both columns to match the "Grand Total" (sum of regular + net),
    # we use total_cr (which is the larger side usually) for both.
    for r in range(2, max_r + 1):
        if is_total_payroll_row(ws, r):
            ws[f"{COL_OUT_DEBT}{r}"].value = float(round(total_cr, 2))
            ws[f"{COL_OUT_CREDIT}{r}"].value = float(round(total_cr, 2))

    # Write net into NET row (Net = Credit Total - Debit Total)
    net_val = float(round(total_cr - total_db, 2))
    for r in range(2, max_r + 1):
        if is_net_row(ws, r):
            # Put net in Db column (G) like you requested before; leave Cr blank
            ws[f"{COL_OUT_DEBT}{r}"].value = None
            ws[f"{COL_OUT_CREDIT}{r}"].value = net_val
            

    # Debug sheet
    dbg_name = "_Calc Debug"
    if dbg_name in wb.sheetnames:
        del wb[dbg_name]
    dbg_ws = wb.create_sheet(dbg_name)

    dbg_ws.append(["Missing Tokens (not found in this workbook's columns):"])
    if missing_tokens:
        for t in sorted(missing_tokens):
            dbg_ws.append([t])
    else:
        dbg_ws.append(["(none)"])

    dbg_ws.append([])
    dbg_ws.append(["Detailed Token Mapping"])
    dbg_ws.append(["Row", "Side", "Department", "Token", "Value", "Sheet", "Column", "Target Date"])
    for row in debug_rows:
        dbg_ws.append(row)

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        TARGET_PAY_DATE = sys.argv[1].strip()
        OUTPUT_XLSX = f"PayrollFilled_{TARGET_PAY_DATE.replace('/','-')}.xlsx"
        print(f"Using target date from command line: {TARGET_PAY_DATE}")
    main()
